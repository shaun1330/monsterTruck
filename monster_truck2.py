import tkinter as tk
import requests
import logging
from sys import exit
from tkinter.ttk import Treeview, Progressbar, Style
from invoice_generator_class import InvoiceGenerator
from receipt_generator_class import ReceiptGenerator
from refund_generator_class import RefundGenerator
from commitee_report import CommiteeReportGenerator
from datetime import datetime
from dateutil.relativedelta import relativedelta
import tkinter.font as font
from tkinter import messagebox
from tkcalendar import DateEntry
from os import startfile
from re import match
from database_connection import MydatabaseConnection as myDb
from email_test import Emailer
from matplotlib.ticker import MultipleLocator
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
from openpyxl import Workbook
from json import loads


version = 'v2.4.0'

button_pressed_color = '#f15000'
button_hover_color = '#1946eb'
button_hover_text_color = 'white'



import configparser

config = configparser.ConfigParser()

config.read('./config/config.ini')

EMAIL_ADDRESS = config['EMAIL']['ADDRESS']
EMAIL_PASSWORD = config['EMAIL']['PASSWORD']
EMAIL_HOST = config['EMAIL']['HOST']
EMAIL_PORT = config['EMAIL']['PORT']

DATABASE_USER = config['DATABASE']['USER']
DATEBASE_PASSWORD = config['DATABASE']['PASSWORD']
DATABASE_HOST = config['DATABASE']['HOST']
DATABASE_NAME = config['DATABASE']['NAME']

FOLDER = config['APP']['FOLDER_PATH']
print(FOLDER)

TOKEN = config['OAUTH']['TOKEN']

scope = 'gmail.send'
redirect_uri = config['OAUTH']['REDIRECT_URIS']
CLIENT_ID = config['OAUTH']['CLIENT_ID']



# custom message box for abort retry and ignore when sending emails
def abortretryignore(title=None, message=None, **options):
    _show = messagebox._show
    s = _show(title, message, messagebox.WARNING, messagebox.ABORTRETRYIGNORE, **options)
    # s might be a Tcl index object, so convert it to a string
    s = str(s)
    if s == messagebox.IGNORE:
        return None
    elif s == messagebox.RETRY:
        return True
    else:
        return False


def on_hover(event):
    event.widget["bg"] = button_hover_color
    event.widget["fg"] = button_hover_text_color
    event.widget["relief"] = tk.RAISED


def off_hover(event):
    event.widget["bg"] = "SystemButtonFace"
    event.widget['fg'] = 'black'
    event.widget['relief'] = tk.RAISED


def activity_log(func):
    def log_wrapper(*args, **kwargs):
        logger.debug(f'Called function {func.__qualname__} with args [{"".join([str(arg) for arg in args])}]', )
        print(f'Called function {func.__qualname__}')
        f = func(*args, **kwargs)
        return f
    return log_wrapper


def check_if_current(startup=True):
    url = 'http://shaunrsimons.com/updates/current_version.txt'
    r = requests.get(url=url)
    request_string = loads(r.text)
    new_version = request_string['version']
    notes = request_string['notes']
    if r.status_code == requests.codes.ok:
        if new_version != version:
            print(f'{new_version} is available for download.')
            top = tk.Toplevel()
            top.title('New Version Available')
            top.attributes('-topmost', True)
            x = top.winfo_screenwidth()
            y = top.winfo_screenheight()
            a = 0.4
            b = 0.4
            top.geometry(str(round(x * a)) +
                         'x'
                         + str(round(y * b)) +
                         '+'
                         + str(round((-a * x + x) / 2)) +
                         '+'
                         + str(round((-b * y + y) / 2)))
            update_label = tk.Label(top, text='Update Available', font=('Courier,12'))
            update_label.pack()
            msg = tk.Text(top, wrap=tk.WORD, height=15, yscrollcommand=set())
            msg.insert('1.0', notes)
            msg.insert('2.0', '\nTo update, close out of Monster Truck and run updater.exe.\n\n')
            msg.configure(state=tk.DISABLED)
            msg.pack(pady=(10,20))
            button = tk.Button(top, text="Dismiss", command=top.destroy,
                               activebackground = button_pressed_color)
            button.bind('<Enter>', on_hover)
            button.bind('<Leave>', off_hover)
            button.pack()


        else:
            if startup:
                print(f'{version} is up to date.')
                pass
            else:
                messagebox.showinfo('Up to date', f'Monster Truck {version} is the most up to date version.')
    else:
        messagebox.showwarning('Update server offline',
                               'Could not connected to update server. Contact developer or try again later.')


def input_check(cash_amount, transfer_amount):
    if len(cash_amount) == 0:
        cash_amount = '0'
    if len(transfer_amount) == 0:
        transfer_amount = '0'

    tests = {'Cash value must be numeric':True,
             'Cash value must be greater than zero':True,
             'Transfer value must be numeric':True,
             'Transfer value must be greater than zero':True}

    if match(r'^-?\d+\.?(\d+)?$', cash_amount) is None:
        tests['Cash value must be numeric'] = False
    elif float(cash_amount) < 0:
        tests['Cash value must be greater than zero'] = False

    if match(r'^-?\d+\.?(\d+)?$', transfer_amount) is None:
        tests['Transfer value must be numeric'] = False
    elif float(transfer_amount) < 0:
        tests['Transfer value must be greater than zero'] = False

    error_string = ''


    if all([value == True for value in tests.values()]):
        return True
    else:

        error_string += 'The following fields are invalid:\n\n'
        for key, value in tests.items():
            if value == False:
                error_string += "-" + key + '\n'
        return error_string


def log_unhandled_exception(type, value, traceback):
    logger.exception('Unhandled Exception', exc_info=(type, value, traceback))
    width, height = 100, 100
    top = tk.Toplevel()
    top.title('Critical Error')
    top.attributes('-topmost', True)

    x = top.winfo_screenwidth()
    y = top.winfo_screenheight()
    a = 0.20
    b = 0.35
    top.geometry(str(round(x * a)) +
                  'x'
                  + str(round(y * b)) +
                  '+'
                  + str(round((-a * x + x) / 2)) +
                  '+'
                  + str(round((-b * y + y) / 2)))

    canvas = tk.Canvas(top, width=width, height=height)
    canvas.pack()

    msg = tk.Message(top, text='An unknown error has occurred.\n Contact Developer. \n\n'
                               f'{type, value}')

    icon = tk.PhotoImage(file='./config/error_icon_sml.png')
    canvas.image = icon
    canvas.create_image(50,50,image=canvas.image)

    msg.pack()
    button = tk.Button(top, text="Dismiss", command=top.destroy,
                       activebackground = button_pressed_color)
    button.bind('<Enter>', on_hover)
    button.bind('<Leave>', off_hover)
    button.pack()


def invoice_table(window, row=0, column=0, columnspan=1, rowspan=1, pady=0, padx=0):
    invoice_font = font.Font(family='Courier', size=10)
    window.invoice_table = Treeview(window, height=10)
    window.invoice_table["columns"] = ('item_desc', 'unit_price', 'qty', 'subtotal')
    window.invoice_table.column('#0', width=invoice_font.measure("Item Code"), stretch=tk.NO)
    window.invoice_table.column('item_desc', width=invoice_font.measure("Annual Membership Renewal Fee"), stretch=tk.NO)
    window.invoice_table.column('unit_price', width=invoice_font.measure("Unit Price"), stretch=tk.NO)
    window.invoice_table.column('qty', width=invoice_font.measure(" QTY "), stretch=tk.NO)
    window.invoice_table.column('subtotal', width=invoice_font.measure("Sub Total"), stretch=tk.NO)

    window.invoice_table.heading('#0', text="Item Code")
    window.invoice_table.heading('item_desc', text="Item Description")
    window.invoice_table.heading('unit_price', text="Unit Price")
    window.invoice_table.heading('qty', text="QTY")
    window.invoice_table.heading('subtotal', text="Sub Total")

    window.invoice_table.grid(row=row, column=column, rowspan=rowspan, columnspan=columnspan, padx=padx, pady=pady)


def parse_settings():
    setting = open('./config/settings.txt', 'r')
    settings = setting.read()
    settings = settings.splitlines()
    settings = [i.split(':') for i in settings]
    settings = dict(settings)
    email_address = settings['email']
    email_host = settings['email_host']
    email_port = settings['smtp_port']
    email_password = settings['email_password']
    database_user = settings['db_user']
    database_password = settings['db_password']
    host = settings['host']
    database_name = settings['db']
    setting.close()
    return email_address, email_host, email_port, email_password, database_user, database_password, host, database_name


def connect_database(database_user, database_password, host, database_name):
    databaseConnection = myDb(database_user, database_password, host, database_name)  # initialise db connection
    return databaseConnection


def not_connected_message(window):
    messagebox.showerror('Database Connection', 'Not connected to database. Add databases credentials to settings.txt',
                         parent=window)


def on_closing():
    top = tk.Toplevel()
    top.title('Exit')
    top.attributes('-topmost', True)
    top.grid_rowconfigure(0, weight=1)
    top.grid_rowconfigure(1, weight=1)
    top.grid_columnconfigure(0, weight=1)
    top.grid_columnconfigure(1, weight=1)
    x = top.winfo_screenwidth()
    y = top.winfo_screenheight()
    a = 0.17
    b = 0.13
    top.geometry(str(round(x * a)) +
                 'x'
                 + str(round(y * b)) +
                 '+'
                 + str(round((-a * x + x) / 2)) +
                 '+'
                 + str(round((-b * y + y) / 2)))

    msg = tk.Label(top, text='Are you sure you want to exit?')
    msg.grid(row=0, column=0, columnspan=2)

    cancelButton = tk.Button(top, text='Cancel', command=top.destroy,
                             activebackground = button_pressed_color)
    cancelButton.grid(row=1, column=0)
    cancelButton.bind('<Enter>', on_hover)
    cancelButton.bind('<Leave>', off_hover)

    okButton = tk.Button(top, text="Ok", command=exit,
                         activebackground = button_pressed_color)
    okButton.grid(row=1, column=1)
    okButton.bind('<Enter>', on_hover)
    okButton.bind('<Leave>', off_hover)


class App(tk.Tk):
    def __init__(self,  email_address, email_password, email_host, email_port, connection, connection_error, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)
        print('Loading App...')
        s = Style()
        s.theme_use('clam')
        self.report_callback_exception = log_unhandled_exception
        self.connection = connection

        self.menubar = tk.Menu(self)

        self.file_menu = tk.Menu(self.menubar, tearoff=0)
        self.file_menu.add_command(label='Check for Update', command=lambda: check_if_current(startup=False))
        self.file_menu.add_command(label='Exit', command=on_closing)
        self.menubar.add_cascade(label='File', menu=self.file_menu)

        self.edit_menu = tk.Menu(self.menubar, tearoff=0)
        self.edit_menu.add_command(label='Edit Categories', command=lambda: EditCategories(self.connection, MainMenu))
        self.edit_menu.add_command(label='Edit Bank Details', command=lambda: BankDetails(self.connection, MainMenu))
        self.menubar.add_cascade(label='Edit', menu=self.edit_menu)

        self.configure(background='black', menu=self.menubar)

        self.screen_height = self.winfo_screenheight()
        self.screen_width = self.winfo_screenwidth()
        self.title(f"Monster Truck {version}")
        self.geometry(str(self.winfo_screenwidth())+'x'+str(self.winfo_screenheight()))
        self.state('zoomed')
        container = tk.Frame(self)
        container.pack(side="top", fill="both", expand=True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.frames = {}
        for F in (MainMenu, Members):
            print(f'Loading {F.__name__}')
            if F == MainMenu:
                frame = F(container, self, connection=self.connection, email_address=email_address,
                          email_password=email_password,
                          email_host=email_host,
                          email_port=email_port)

                self.frames[F] = frame
                frame.grid(row=0, column=0, sticky='NSEW')
            else:
                frame = F(container, self, connection=self.connection)
                self.frames[F] = frame
                frame.grid(row=0, column=0, sticky='NSEW')
            print(f'\t{F.__name__} Loaded')
        self.show_frame(MainMenu)
        if self.connection == None and connection_error == 1:
            messagebox.showwarning('Database connection', 'A connection could not be made with the database. '
                                                          'Check that credentials are correct.')
        elif self.connection == None and connection_error == 0:
            messagebox.showinfo('Database connection', 'A connection could not be made with the database. '
                                                       'Database credentials are missing. Check settings.txt.')
        elif self.connection != None:
            self.check_database_connection()
        # check_if_current()

    def show_frame(self, page):
        frame = self.frames[page]
        frame.tkraise()

    def check_database_connection(self):
        time_now = datetime.now().strftime('%H:%M:%S')
        connected = self.connection.is_connect()
        if connected:
            print(time_now + ': Connected')
            self.after(1000 * 60 * 5, self.check_database_connection)  # check every 5 minutes
        else:
            messagebox.showwarning('Database connection lost',
                                   'The connection to the database has been lost. Restart Monster Truck.')


class MainMenu(tk.Frame):
    def __init__(self, parent, controller, connection, email_address, email_password, email_host, email_port):
        tk.Frame.__init__(self, parent)
        self.connect = connection
        self.email_address = email_address
        self.email_password = email_password
        self.email_host = email_host
        self.email_port = email_port
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure(2, weight=1)
        self.grid_columnconfigure(3, weight=1)
        self.grid_columnconfigure(4, weight=1)
        self.grid_columnconfigure(5, weight=1)
        self.grid_columnconfigure(6, weight=1)
        self.grid_columnconfigure(7, weight=1)
        self.grid_columnconfigure(8, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self.grid_rowconfigure(2, weight=1)
        self.grid_rowconfigure(3, weight=1)
        self.grid_rowconfigure(4, weight=1)
        self.grid_rowconfigure(5, weight=1)
        self.grid_rowconfigure(6, weight=1)
        self.grid_rowconfigure(7, weight=1)
        self.grid_rowconfigure(8, weight=1)
        if connection != None:
            self.unsent_invoices = self.connect.query('select '
                                                      'concat(invoice_no,".pdf") as filename, '
                                                      'email '
                                                      'from '
                                                      'invoice '
                                                      'join '
                                                      'members '
                                                      'on '
                                                      'invoice.member_no = members.member_no '
                                                      'where invoice_sent = "No";')
        else:
            self.unsent_invoices = []

        button_font = font.Font(family='Courier', size=15, weight='bold')

        self.member_button = tk.Button(self, text='Members',
                                           command=lambda: controller.show_frame(Members),
                                           activebackground=button_pressed_color)
        self.member_button['font'] = button_font
        self.member_button.grid(row=1, column=0)
        self.member_button.bind('<Enter>', on_hover)
        self.member_button.bind('<Leave>', off_hover)

        self.new_invoice_button = tk.Button(self, text='New Invoice',
                                            command=lambda: InvoiceWindow(connection, self),
                                            activebackground=button_pressed_color)
        self.new_invoice_button['font'] = button_font
        self.new_invoice_button.grid(row=2, column=0)
        self.new_invoice_button.bind('<Enter>', on_hover)
        self.new_invoice_button.bind('<Leave>', off_hover)

        self.new_expense_button = tk.Button(self, text='Add\nExpense',
                                            command=lambda: ExpenseWindow(connection, self),
                                            activebackground=button_pressed_color)
        self.new_expense_button['font'] = button_font
        self.new_expense_button.grid(row=3, column=0)
        self.new_expense_button.bind('<Enter>', on_hover)
        self.new_expense_button.bind('<Leave>', off_hover)

        self.new_income_button = tk.Button(self, text='Add\nIncome',
                                           command=lambda: IncomeWindow(connection, self),
                                           activebackground=button_pressed_color)
        self.new_income_button['font'] = button_font
        self.new_income_button.grid(row=4, column=0)
        self.new_income_button.bind('<Enter>', on_hover)
        self.new_income_button.bind('<Leave>', off_hover)

        self.cash_deposit_button = tk.Button(self, text='Cash Out/\nBank Deposit',
                                             command=lambda: CashTransfers(connection, self),
                                             activebackground=button_pressed_color)
        self.cash_deposit_button['font'] = button_font
        self.cash_deposit_button.grid(row=5, column=0)
        self.cash_deposit_button.bind('<Enter>', on_hover)
        self.cash_deposit_button.bind('<Leave>', off_hover)

        self.auto_invoicer_button = tk.Button(self, text='Auto\nInvoicer',
                                              command=lambda: AutoInvoicing(connection,
                                                                            self, self.email_address,
                                                                            self.email_password,
                                                                            self.email_host,
                                                                            self.email_port),
                                              activebackground=button_pressed_color)
        self.auto_invoicer_button['font'] = button_font
        self.auto_invoicer_button.grid(row=6, column=0)
        self.auto_invoicer_button.bind('<Enter>', on_hover)
        self.auto_invoicer_button.bind('<Leave>', off_hover)

        self.records_button = tk.Button(self, text='Committee\nReport',
                                        command=lambda: ReportPeriod(self.connect, self),
                                        activebackground=button_pressed_color)
        self.records_button['font'] = button_font
        self.records_button.grid(row=7, column=0)
        self.records_button.bind('<Enter>', on_hover)
        self.records_button.bind('<Leave>', off_hover)

        self.send_unsent_button = tk.Button(self, text='Send Unsent\nInvoices',
                                            command=self.send_unsent_invoices,
                                            activebackground=button_pressed_color)
        self.send_unsent_button['font'] = button_font
        self.send_unsent_button.grid(row=8, column=0)
        self.send_unsent_button.bind('<Enter>', on_hover)
        self.send_unsent_button.bind('<Leave>', off_hover)

        invoice_table_label = tk.Label(self, text='Unpaid Invoices')
        invoice_table_label.config(font="Courier, 14")
        invoice_table_label.grid(row=0, column=1, sticky='N', columnspan=2)

        self.invoices_table = Treeview(self, height=20)
        self.invoices_table["columns"] = ('member_name', 'issue_date', 'due_date', 'total', 'invoice_sent')
        self.invoices_table.grid(row=1, column=1, rowspan=5, columnspan=2)

        self.invoices_table.column('#0', width=round(self.winfo_screenwidth() * 0.06), stretch=False)
        self.invoices_table.heading('#0', text="Invoice No")
        self.invoices_table.column('member_name', width=round(self.winfo_screenwidth() * 0.1), stretch=False)
        self.invoices_table.heading('member_name', text="Member Name")
        self.invoices_table.column('issue_date', width=round(self.winfo_screenwidth() * 0.06), stretch=False)
        self.invoices_table.heading('issue_date', text="Issue Date")
        self.invoices_table.column('due_date', width=round(self.winfo_screenwidth() * 0.06), stretch=False)
        self.invoices_table.heading('due_date', text="Due Date")
        self.invoices_table.column('total', width=round(self.winfo_screenwidth() * 0.06), stretch=False)
        self.invoices_table.heading('total', text="Invoice Total")
        self.invoices_table.column('invoice_sent', width=round(self.winfo_screenwidth() * 0.03), stretch=False)
        self.invoices_table.heading('invoice_sent', text="Sent")
        self.invoices_table.bind("<Double-1>", self.on_double_click_invoice)

        history_table_label = tk.Label(self, text='Transactions History')
        history_table_label.config(font="Courier, 14")
        history_table_label.grid(row=0, column=3, sticky='N')
        # button_font = font.Font(family='Courier', size=15, weight='bold')
        self.history_column_font = font.Font(family='Courier', size=10)

        self.history_table = Treeview(self, height=36)
        self.history_table["columns"] = ('code',
                                         'date',
                                         'cash_amount',
                                         'transfer_amount',
                                         'cash_balance',
                                         'bank_balance')
        self.history_table.grid(row=1, column=3, rowspan=8, sticky='w')

        self.history_table.column('#0', width=self.history_column_font.measure('EXPENSE  '),
                                  stretch=False)
        self.history_table.heading('#0', text="Type")
        self.history_table.column('code', width=self.history_column_font.measure('99999'),
                                  stretch=False)
        self.history_table.heading('code', text="Code")
        self.history_table.column('date', width=self.history_column_font.measure('30-22-9999'),
                                  stretch=False, anchor='e')
        self.history_table.heading('date', text='Date')
        self.history_table.column('cash_amount', width=self.history_column_font.measure('Cash Amount'),
                                  stretch=False, anchor='e')
        self.history_table.heading('cash_amount', text="Cash Amount")
        self.history_table.column('transfer_amount', width=self.history_column_font.measure('Transfer Amount'),
                                  stretch=False, anchor='e')
        self.history_table.heading('transfer_amount', text="Transfer Amount")
        self.history_table.column('cash_balance', width=self.history_column_font.measure('Cash Balance'),
                                  stretch=False, anchor='e')
        self.history_table.heading('cash_balance', text="Cash Balance")
        self.history_table.column('bank_balance', width=self.history_column_font.measure('Bank Balance'),
                                  stretch=False, anchor='e')
        self.history_table.heading('bank_balance', text="Bank Balance")
        self.history_table.bind("<Double-1>", self.on_double_click_history)

        if connection != None:
            self.populate_invoice_table()
            self.populate_history_table()

            try:
                self.exporter()
            except PermissionError:
                messagebox.showerror('Members Export Error', '"members_list.xlsx" failed to update.\n'
                                                             'Make sure "members_list.xlsx" is not open.')
                # error when file is open already

        self.update_balance()

    @activity_log
    def exporter(self):
        print(f'\t\t\tExporting member data to members_list.xlsx{"."*10}', end='')
        members_info = self.connect.query('select '
                                                     'member_no, '
                                                     'member_fname, '
                                                     'member_lname, '
                                                     'partner_name, '
                                                     'street_address, '
                                                     'suburb, postcode, state, '
                                                     'home_phone, '
                                                     'mobile_phone, '
                                                     'email '
                                                     'from members where member_no != 1 and member_status = "ACTIVE"')

        workbook = Workbook()
        sheet = workbook.active

        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 15
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 25
        sheet.column_dimensions['F'].width = 20
        sheet.column_dimensions['G'].width = 10
        sheet.column_dimensions['H'].width = 10
        sheet.column_dimensions['I'].width = 15
        sheet.column_dimensions['J'].width = 15
        sheet.column_dimensions['K'].width = 30

        sheet['A1'] = 'Member No'
        sheet['B1'] = 'First Name'
        sheet['C1'] = 'Last Name'
        sheet['D1'] = 'Partner Name'
        sheet['E1'] = 'Street Address'
        sheet['F1'] = 'Suburb'
        sheet['G1'] = 'Postcode'
        sheet['H1'] = 'State'
        sheet['I1'] = 'Home Phone'
        sheet['J1'] = 'Mobile Phone'
        sheet['K1'] = 'Email'

        r = 2
        c = 1
        for member in members_info:
            for element in member:
                cell = sheet.cell(row=r, column=c)
                cell.value = element
                c += 1
            c = 1
            r += 1

        workbook.save(filename='.\config\members_list.xlsx')
        print(" members_list.xlsx saved")

    def send_unsent_invoices(self):
        if self.connect != None:
            self.unsent_invoices = self.connect.query('select '
                                                      'concat(invoice_no,".pdf") as filename, '
                                                      'email, '
                                                      'concat_ws(" ", member_fname, member_lname)'
                                                      'from '
                                                      'invoice '
                                                      'join '
                                                      'members '
                                                      'on '
                                                      'invoice.member_no = members.member_no '
                                                      'where invoice_sent = "No" and '
                                                      'invoice_no not in '
                                                      '(select invoice_no from invoice_receipt);')
            if len(self.unsent_invoices) == 0:
                messagebox.showinfo('Send Invoices', 'There are no unsent invoices')
            else:
                send_yesno = messagebox.askyesno('Send Invoices', 'Are you sure you want to send out unsent invoices?')
                if send_yesno:
                    status = EmailProgress(self.unsent_invoices, self.connect, self, self.email_address,
                                           self.email_password,
                                           self.email_host,
                                           self.email_port)
                    if status.get_error_status() == '1':
                        messagebox.showinfo('Email Invoices', 'All invoices sent out', parent=self)
        else:
            not_connected_message(self)

    def update_tables(self):
        self.invoices_table.delete(*self.invoices_table.get_children())  # delete invoice table rows
        self.history_table.delete(*self.history_table.get_children())  # delete history table rows
        self.populate_invoice_table()  # refill table with updated data
        self.populate_history_table()

    def populate_invoice_table(self):
        unpaid_invoices = self.connect.query('select'
                                             ' invoice.invoice_no,'
                                             ' concat(member_fname," ", member_lname) as member_name,'
                                             ' DATE_FORMAT(invoice_date, "%d-%m-%Y") as issue_date,'
                                             ' DATE_FORMAT(invoice_duedate,"%d-%m-%Y") as due_date,'
                                             ' invoice_total, invoice_sent '
                                             'from invoice join members on invoice.member_no = members.member_no'
                                             ' where invoice_no not in (select invoice_no from invoice_receipt) '
                                             'order by invoice_no desc;')

        for row in unpaid_invoices:
            self.invoices_table.insert('', 'end', text=row[0], values=row[1:])

    def populate_history_table(self):
        self.create_temp_table()
        self.transaction_history = self.connect.query('select '
                                                 'n, date_format(payment_datetime,"%d-%m-%Y"), '
                                                 'cash_amount, '
                                                 'transfer_amount, '
                                                 'cash_balance, '
                                                 'bank_balance '
                                                 'from payment_history '
                                                 'where n != 30000 '
                                                 'order by payment_datetime desc, n desc;')
        for row in self.transaction_history:
            if row[0] > 30000 and row[0] < 40000:
                self.history_table.insert('', 'end', text='Income', values=row)
            elif 40000 < row[0] < 50000:
                self.history_table.insert('', 'end', text='Invoice', values=row)
            elif row[0] > 50000 and row[0] < 70000:
                if row[2] > 0:
                    self.history_table.insert('', 'end', text='Cash out', values=row)
                elif row[2] < 0:
                    self.history_table.insert('', 'end', text='Deposit', values=row)
            elif row[0] > 70000 and row[0] < 90000:
                self.history_table.insert('', 'end', text='Expense', values=row)
            elif row[0] > 90000:
                self.history_table.insert('', 'end', text='Refund', value=row)
        self.update_balance()

    def create_temp_table(self):
        try:
            self.connect.insert('set @cashSum := 0;')
            self.connect.insert('set @bankSum := 0;')
            print(f'\t\tDropping payment_history{"."*10}', end='')
            self.connect.insert('drop table if exists payment_history;')
            print('Dropped')
            print(f'\t\tCreating new payment_history{"."*10}', end='')
            self.connect.insert('create table payment_history '
                                'select '
                                'n, payment_datetime, '
                                'cash_amount, '
                                'transfer_amount, '
                                '(@cashSum := @cashSum + cash_amount) as cash_balance, '
                                '(@bankSum := @bankSum + transfer_amount) as bank_balance '
                                'from '
                                '((select invoice_receipt_no as n, '
                                'payment_datetime,'
                                ' cash_amount, '
                                'transfer_amount'
                                ' from invoice_receipt where invoice_receipt_no != 40000) '
                                'union all'
                                ' (select expense_receipt_no as n, '
                                'payment_datetime, '
                                'cash_amount, '
                                'transfer_amount '
                                'from expense_receipt where expense_receipt_no != 70000)'
                                'union all'
                                ' (select transfer_no as n, '
                                'payment_datetime, '
                                'cash_amount, '
                                'transfer_amount '
                                'from transfer where transfer_no != 50000) '
                                'union all'
                                ' (select income_receipt_no as n, '
                                'payment_datetime, '
                                'cash_amount, '
                                'transfer_amount '
                                'from income_receipt) '
                                'union all'
                                ' (select refund_no as n, '
                                'payment_datetime, '
                                'cash_amount*-1, '
                                'transfer_amount*-1 '
                                'from refunds)) as hist '
                                'order by payment_datetime;')
            self.connect.commit()
            print('Done')
        except Exception as e:
            self.connect.rollback()
            logger.exception(e)
            messagebox.showwarning('Unknown Error',
                                   f'An unknown error occurred and has been logged. Report to developer.',
                                   parent=self)

    @activity_log
    def on_double_click_invoice(self, a):
        if self.connect != None:
            item = self.invoices_table.selection()
            invoice_no = self.invoices_table.item(item, 'text')
            if invoice_no != '':
                Receipt_window(invoice_no, self.connect, self, self.email_address, self.email_password, self.email_host, self.email_port)

    @activity_log
    def on_double_click_history(self, a):
        if self.connect != None:
            item = self.history_table.selection()
            receipt_no = self.history_table.item(item, 'value')
            receipt_type = self.history_table.item(item, 'text')
            if receipt_type != '':
                HistoryWindow(receipt_no[0], self.connect, self, receipt_type, self.email_address, self.email_password, self.email_host, self.email_port)

    def money_in_out_labels(self):
        if self.connect != None:
            self.balances = self.transaction_history[0][-2:]
            self.month_in = self.connect.query('select '
                                                       'sum(cash_amount)+sum(transfer_amount) '
                                                       'from '
                                                       '(\n'
                                                       '(select invoice_no as n, '
                                                       'payment_datetime, '
                                                       'cash_amount, '
                                                       'transfer_amount '
                                                       'from invoice_receipt)'
                                                       '\n'
                                                       'union all'
                                                       '\n'
                                                       '(select expense_receipt_no as n, '
                                                       'payment_datetime, '
                                                       'cash_amount, '
                                                       'transfer_amount '
                                                       'from expense_receipt)'
                                                       '\n'
                                                       'union all'
                                                       '\n'
                                                       '(select income_receipt_no as n, '
                                                       'payment_datetime, '
                                                       'cash_amount, '
                                                       'transfer_amount '
                                                       'from income_receipt)) '
                                                       'as hist\n'
                                                       'where '
                                                       '(payment_datetime '
                                                       'between date_format(now() , "%Y-%m-01") and NOW())'
                                                       ' and '
                                                       '(cash_amount > 0 or transfer_amount > 0)\n'
                                                       'order by payment_datetime desc;')
            self.month_out = self.connect.query('select '
                                                       'sum(cash_amount)+sum(transfer_amount) '
                                                       'from '
                                                       '(\n'
                                                       '(select invoice_no as n, '
                                                       'payment_datetime, '
                                                       'cash_amount, '
                                                       'transfer_amount '
                                                       'from invoice_receipt)'
                                                       '\n'
                                                       'union all'
                                                       '\n'
                                                       '(select expense_receipt_no as n, '
                                                       'payment_datetime, '
                                                       'cash_amount, '
                                                       'transfer_amount '
                                                       'from expense_receipt)'
                                                       '\n'
                                                       'union all'
                                                       '\n'
                                                       '(select income_receipt_no as n, '
                                                       'payment_datetime, '
                                                       'cash_amount, '
                                                       'transfer_amount '
                                                       'from income_receipt)) '
                                                       'as hist\n'
                                                       'where '
                                                       '(payment_datetime '
                                                       'between date_format(now() , "%Y-%m-01") and NOW())'
                                                       ' and '
                                                       '(cash_amount < 0 or transfer_amount < 0)\n'
                                                       'order by payment_datetime desc;')
            self.month_in = self.month_in[0][0]
            self.month_out = self.month_out[0][0]
        else:
            self.month_in = 0
            self.month_out = 0
            self.balances = (0,0)

        self.cash_balance_label = tk.Label(self,
                                           text=f'Cash Balance: ${self.balances[0]}\n'
                                                f'Bank Balance: ${self.balances[1]}',
                                           font='courier 15 bold')
        self.cash_balance_label.grid(row=6, column=1)
        current_month = datetime.today().month
        if current_month == 1:
            self.current_month = 'January'
        elif current_month == 2:
            self.current_month = 'February'
        elif current_month == 3:
            self.current_month = 'March'
        elif current_month == 4:
            self.current_month = 'April'
        elif current_month == 5:
            self.current_month = 'May'
        elif current_month == 6:
            self.current_month = 'June'
        elif current_month == 7:
            self.current_month = 'July'
        elif current_month == 8:
            self.current_month = 'August'
        elif current_month == 9:
            self.current_month = 'September'
        elif current_month == 10:
            self.current_month = 'October'
        elif current_month == 11:
            self.current_month = 'November'
        elif current_month == 12:
            self.current_month = 'December'

        self.money_in_out_label = tk.Label(self,
                                           text=f'{self.current_month} Money In: ${self.month_in}\n{self.current_month} Money Out: ${self.month_out}',
                                           font='courier 15 bold')
        self.money_in_out_label.grid(row=6, column=2)
        if self.month_in is None:
            self.month_in = 0
        if self.month_out is None:
            self.month_out = 0
        self.money_in_out_label.configure(text=f'{self.current_month} Money In: ${self.month_in}\n{self.current_month} Money Out: ${self.month_out}',
                                                  font='courier 15 bold')

    def end_of_month_balance_graph(self):
        if self.connect != None:
            self.end_of_month_balance = self.connect.query(
                'select date_format(x.payment_datetime, "%Y-%m") , x.cash_balance, x.bank_balance '
                'from payment_history as x join ('
                '   select max(payment_datetime) as payment_datetime from payment_history as y '
                '   group by year(y.payment_datetime), month(y.payment_datetime))'
                '   as z using (payment_datetime) where n != 30000 group by payment_datetime '
                '   order by payment_datetime desc')
            self.clean_monthly_balances()
        else:
            self.cleaned_balances = [(datetime.today() - relativedelta(months=i),0,0) for i in range(0,6)]



        months = [row[0] for row in self.cleaned_balances]
        months.reverse()
        cash_balances = [row[1] for row in self.cleaned_balances]
        cash_balances.reverse()
        bank_balances = [row[2] for row in self.cleaned_balances]
        bank_balances.reverse()
        total_balances = [row[1] + row[2] for row in self.cleaned_balances]
        total_balances.reverse()

        f = Figure(figsize=(6, 2))
        f.suptitle('Monthly Closing Balances')
        f.set_facecolor((0.94, 0.94, 0.93))
        a = f.add_subplot(111, ylim=(0, float(max(total_balances)) + 100))
        a.bar(months, total_balances, label='Total', color='#1946eb')
        a.plot(months, cash_balances, color='#f15000', label='Cash', marker='.')
        a.plot(months, bank_balances, color='black', label='Bank', marker='.')

        a.set_ylabel('Dollars ($)')
        major_spacing = len(months) // 5
        print(major_spacing)
        if major_spacing > 1:
            a.xaxis.set_major_locator(MultipleLocator(major_spacing))
            a.xaxis.set_minor_locator(MultipleLocator(1))

        a.tick_params(axis='x', labelsize=7)
        a.tick_params(axis='x', which='major', length=4)
        a.tick_params(axis='x', which='minor', length=2)
        a.tick_params(axis='y', labelsize=7)
        a.legend(ncol=3, fontsize=7)

        a.set_facecolor((0.94, 0.94, 0.93))

        canvas = FigureCanvasTkAgg(f, self)
        canvas.get_tk_widget().grid(row=7, column=1, rowspan=2, columnspan=2, sticky='n')
        #
        # plt.figure(dpi=200)
        # plt.bar(months, balances, color='grey')
        # plt.xticks(rotation=30)
        # plt.savefig('./config/bar_closing')

    def clean_monthly_balances(self):
        self.cleaned_balances = []
        for i in range(len(self.end_of_month_balance) - 1):
            row = self.end_of_month_balance[i]
            current_row_month = datetime.strptime(row[0], "%Y-%m")
            next_row_month = datetime.strptime(self.end_of_month_balance[i + 1][0], "%Y-%m")
            while current_row_month != next_row_month:
                self.cleaned_balances.append((current_row_month.strftime("%Y-%m"), row[1], row[2]))
                current_row_month = current_row_month - relativedelta(months=1)
        self.cleaned_balances.append(self.end_of_month_balance[-1])

    def update_balance(self):
        self.money_in_out_labels()
        self.end_of_month_balance_graph()


class Members(tk.Frame):
    def __init__(self, parent, controller, connection):
        tk.Frame.__init__(self, parent)
        self.databaseConnection = connection
        self.screen_height = self.winfo_screenheight()
        self.screen_width = self.winfo_screenwidth()

        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure(2, weight=1)
        self.grid_columnconfigure(3, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self.grid_rowconfigure(2, weight=1)

        if self.databaseConnection != None:
            self.active_members_table = self.databaseConnection.query('select '
                                                      'member_no, '
                                                      'member_fname, '
                                                      'member_lname, '
                                                      'partner_name, '
                                                      'street_address, '
                                                      'suburb, postcode, '
                                                      'home_phone, '
                                                      'mobile_phone, '
                                                      'email, '
                                                      'member_status '
                                                      'from members where member_no != 1 and '
                                                      'member_status = "ACTIVE";')
        else:
            self.active_members_table = []


        self.member_label = tk.Label(self, text='Members Menu', font='courier 40 bold')
        self.member_label.grid(row=0, columnspan=4)

        self.table = Treeview(self, height=20)
        self.table.bind("<Double-1>", self.on_double_click)
        self.table["columns"] = \
            (
            'Fname',
            'Lname',
            'partner',
            'address',
            'suburb',
            'postcode',
            'homePhone',
            'mobile',
            'email',
            'status'
            )

        self.table.column('#0', width=round(self.screen_width*0.05), minwidth=50, stretch=tk.NO)
        self.table.column('Fname', width=round(self.screen_width*0.07), minwidth=50, stretch=tk.NO)
        self.table.column('Lname', width=round(self.screen_width*0.07), minwidth=50, stretch=tk.NO)
        self.table.column('partner', width=round(self.screen_width * 0.10), minwidth=50, stretch=tk.NO)
        self.table.column('address', width=round(self.screen_width * 0.12), minwidth=50, stretch=tk.NO)
        self.table.column('suburb', width=round(self.screen_width * 0.10), minwidth=50, stretch=tk.NO)
        self.table.column('postcode', width=round(self.screen_width * 0.05), minwidth=50, stretch=tk.NO)
        self.table.column('homePhone', width=round(self.screen_width * 0.07), minwidth=50, stretch=tk.NO)
        self.table.column('mobile', width=round(self.screen_width*0.07), minwidth=50, stretch=tk.NO)
        self.table.column('email', width=round(self.screen_width*0.15), minwidth=50, stretch=tk.NO)
        self.table.column('status', width=round(self.screen_width * 0.07), minwidth=50, stretch=tk.NO)

        self.table.heading('#0', text="ID")
        self.table.heading('Fname', text='First Name')
        self.table.heading('Lname', text='Last Name')
        self.table.heading('partner', text='Partner Name')
        self.table.heading('address', text='Street Address')
        self.table.heading('suburb', text='Suburb')
        self.table.heading('postcode', text='Postcode')
        self.table.heading('homePhone', text='Home Phone')
        self.table.heading('mobile', text='Mobile Number')
        self.table.heading('email', text='Email')
        self.table.heading('status', text='Member Status')
        self.table.grid(row=1, columnspan=4, padx=round(self.screen_width*0.02))

        for row in self.active_members_table:
            self.table.insert('', 'end', text=row[0], values=row[1:])

        button_font = font.Font(family='Courier', size=20, weight='bold')

        self.new_member_button = tk.Button(self, text='New Member', command=lambda: NewMemberPage(connection, self),
                                           activebackground=button_pressed_color)
        self.new_member_button['font'] = button_font
        self.new_member_button.grid(row=2, column=1, sticky='N', pady=5)
        self.new_member_button.bind('<Enter>', on_hover)
        self.new_member_button.bind('<Leave>', off_hover)

        self.main_menu_button = tk.Button(self, text='Main Menu', command=lambda: controller.show_frame(MainMenu),
                                          activebackground=button_pressed_color)
        self.main_menu_button['font'] = button_font
        self.main_menu_button.grid(row=2, column=0, sticky='N', pady=5)
        self.main_menu_button.bind('<Enter>', on_hover)
        self.main_menu_button.bind('<Leave>', off_hover)


        self.export_button = tk.Button(self, text='Export Members Info', command=self.exporter,
                                       activebackground=button_pressed_color)
        self.export_button['font'] = button_font
        self.export_button.grid(row=2, column=2, sticky='N', pady=5)
        self.export_button.bind('<Enter>', on_hover)
        self.export_button.bind('<Leave>', off_hover)


        self.toggle_active_button = tk.Button(self, text='Show Inactive Members', command=self.toggle_active,
                                              activebackground=button_pressed_color)
        self.toggle_active_button['font'] = button_font
        self.toggle_active_button.grid(row=2, column=3, sticky='N', pady=5)
        self.toggle_active_button.bind('<Enter>', on_hover)
        self.toggle_active_button.bind('<Leave>', off_hover)

        self.showing_active = True


    @activity_log
    def toggle_active(self):
        if self.databaseConnection != None:
            if self.showing_active:
                self.full_members_table = self.databaseConnection.query('select '
                                                                        'member_no, '
                                                                        'member_fname, '
                                                                        'member_lname, '
                                                                        'partner_name, '
                                                                        'street_address, '
                                                                        'suburb, postcode, '
                                                                        'home_phone, '
                                                                        'mobile_phone, '
                                                                        'email, '
                                                                        'member_status '
                                                                        'from members where member_no != 1 ;')
                self.table.delete(*self.table.get_children())
                for row in self.full_members_table:
                    self.table.insert('', 'end', text=row[0], values=row[1:])
                self.showing_active = False
                self.toggle_active_button['text'] = "Hide Inactive Members"
            else:
                self.table.delete(*self.table.get_children())
                self.active_members_table = self.databaseConnection.query('select '
                                                                          'member_no, '
                                                                          'member_fname, '
                                                                          'member_lname, '
                                                                          'partner_name, '
                                                                          'street_address, '
                                                                          'suburb, postcode, '
                                                                          'home_phone, '
                                                                          'mobile_phone, '
                                                                          'email, '
                                                                          'member_status '
                                                                          'from members where member_no != 1 and '
                                                                          'member_status = "ACTIVE";')
                for row in self.active_members_table:
                    self.table.insert('', 'end', text=row[0], values=row[1:])
                self.showing_active = True
                self.toggle_active_button['text'] = "Show Inactive Members"
        else:
            not_connected_message(self)

    @activity_log
    def exporter(self):
        if self.databaseConnection != None:
            members_info = self.databaseConnection.query('select '
                                                         'member_no, '
                                                         'member_fname, '
                                                         'member_lname, '
                                                         'partner_name, '
                                                         'street_address, '
                                                         'suburb, postcode, state, '
                                                         'home_phone, '
                                                         'mobile_phone, '
                                                         'email '
                                                         'from members where member_no != 1 and member_status = "ACTIVE"')

            workbook = Workbook()
            sheet = workbook.active

            sheet.column_dimensions['A'].width = 10
            sheet.column_dimensions['B'].width = 15
            sheet.column_dimensions['C'].width = 15
            sheet.column_dimensions['D'].width = 20
            sheet.column_dimensions['E'].width = 25
            sheet.column_dimensions['F'].width = 20
            sheet.column_dimensions['G'].width = 10
            sheet.column_dimensions['H'].width = 10
            sheet.column_dimensions['I'].width = 15
            sheet.column_dimensions['J'].width = 15
            sheet.column_dimensions['K'].width = 30

            sheet['A1'] = 'Member No'
            sheet['B1'] = 'First Name'
            sheet['C1'] = 'Last Name'
            sheet['D1'] = 'Partner Name'
            sheet['E1'] = 'Street Address'
            sheet['F1'] = 'Suburb'
            sheet['G1'] = 'Postcode'
            sheet['H1'] = 'State'
            sheet['I1'] = 'Home Phone'
            sheet['J1'] = 'Mobile Phone'
            sheet['K1'] = 'Email'

            r = 2
            c = 1
            for member in members_info:
                for element in member:
                    cell = sheet.cell(row=r, column=c)
                    cell.value = element
                    c += 1
                c = 1
                r += 1

            workbook.save(filename='.\config\members_list.xlsx')
            messagebox.showinfo('Update Success', 'member_list.xlsx has been successfully updated', parent=self)
        else:
            not_connected_message(self)

    @activity_log
    def on_double_click(self, a):
        item = self.table.selection()
        if len(item) != 0:
            member_no = self.table.item(item, 'text')
            EditMember(member_no, self.databaseConnection, self)

    def update_window(self):
        # self.databaseConnection.reconnect()
        self.table.delete(*self.table.get_children())
        if self.showing_active:
            self.active_members_table = self.databaseConnection.query('select '
                                                          'member_no, '
                                                          'member_fname, '
                                                          'member_lname, '
                                                          'partner_name, '
                                                          'street_address, '
                                                          'suburb, postcode, '
                                                          'home_phone, '
                                                          'mobile_phone, '
                                                          'email, '
                                                          'member_status '
                                                          'from members where member_no != 1 and '
                                                          'member_status = "ACTIVE";')
            for row in self.active_members_table:
                self.table.insert('', 'end', text=row[0], values=row[1:])
        else:
            self.full_members_table = self.databaseConnection.query('select '
                                                          'member_no, '
                                                          'member_fname, '
                                                          'member_lname, '
                                                          'partner_name, '
                                                          'street_address, '
                                                          'suburb, postcode, '
                                                          'home_phone, '
                                                          'mobile_phone, '
                                                          'email, '
                                                          'member_status '
                                                          'from members where member_no != 1;')
            for row in self.full_members_table:
                self.table.insert('', 'end', text=row[0], values=row[1:])


class EditMember(tk.Tk):
    def __init__(self, member_no, connection, member_page, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)
        self.report_callback_exception = log_unhandled_exception
        self.member_page = member_page
        self.databaseConnection = connection
        self.member_no = member_no
        values = self.databaseConnection.query('select '
                                               'member_no, '
                                               'member_fname, '
                                               'member_lname, '
                                               'partner_name, '
                                               'street_address, '
                                               'suburb, '
                                               'state, '
                                               'postcode, '
                                               'home_phone, '
                                               'mobile_phone, '
                                               'email, '
                                               'member_status '
                                               'from '
                                               'members '
                                               'where '
                                               f'member_no = {member_no}')
        values = values[0]
        height = self.winfo_screenheight()
        width = self.winfo_screenwidth()
        self.geometry(str(round(width * 0.3)) + 'x' + str(round(height * 0.6)))
        self.fname_var = tk.StringVar(self)
        self.fname_var.set(values[1])
        self.lname_var = tk.StringVar(self)
        self.lname_var.set(values[2])
        self.partner_var = tk.StringVar(self)
        self.partner_var.set(values[3])
        self.address_var = tk.StringVar(self)
        self.address_var.set(values[4])
        self.suburb_var = tk.StringVar(self)
        self.suburb_var.set(values[5])
        self.state_var = tk.StringVar(self)
        self.state_var.set(values[6])
        self.postcode_var = tk.StringVar(self)
        self.postcode_var.set(values[7])
        self.home_phone_var = tk.StringVar(self)
        self.home_phone_var.set(values[8])
        self.mobile_phone_var = tk.StringVar(self)
        self.mobile_phone_var.set(values[9])
        self.email_var = tk.StringVar(self)
        self.email_var.set(values[10])
        self.member_status_var = tk.StringVar(self)
        self.member_status_var.set(values[11])

        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self.grid_rowconfigure(2, weight=1)
        self.grid_rowconfigure(3, weight=1)
        self.grid_rowconfigure(4, weight=1)
        self.grid_rowconfigure(5, weight=1)
        self.grid_rowconfigure(6, weight=1)
        self.grid_rowconfigure(7, weight=1)
        self.grid_rowconfigure(8, weight=1)
        self.grid_rowconfigure(9, weight=1)
        self.grid_rowconfigure(10, weight=1)
        self.grid_rowconfigure(11, weight=1)
        self.grid_rowconfigure(12, weight=1)
        self.grid_rowconfigure(13, weight=1)

        self.title("Edit Member")
        titleLabel = tk.Label(self, text='Edit Member')
        titleLabel.grid(row=0, column=0)

        fnameLabel = tk.Label(self, text='First Name:')
        fnameLabel.grid(column=0, row=2)
        fnameEntry = tk.Entry(self, textvariable=self.fname_var, width=round(self.winfo_screenwidth() * 0.03))
        fnameEntry.grid(column=1, row=2, sticky='W', padx=10, pady=5)

        lnameLabel = tk.Label(self, text='Last Name:')
        lnameLabel.grid(column=0, row=3)
        lnameEntry = tk.Entry(self, textvariable=self.lname_var, width=round(self.winfo_screenwidth() * 0.03))
        lnameEntry.grid(column=1, row=3, sticky='W', padx=10, pady=5)

        partnerLabel = tk.Label(self, text='Partner Full name:')
        partnerLabel.grid(column=0, row=4)
        partnerEntry = tk.Entry(self, textvariable=self.partner_var, width=round(self.winfo_screenwidth() * 0.03))
        partnerEntry.grid(column=1, row=4, sticky='W', padx=10, pady=5)

        addressLabel = tk.Label(self, text='Street Address:')
        addressLabel.grid(column=0, row=5)
        addressEntry = tk.Entry(self, textvariable=self.address_var, width=round(self.winfo_screenwidth() * 0.03))
        addressEntry.grid(column=1, row=5, sticky='W', padx=10, pady=5)

        suburbLabel = tk.Label(self, text='Suburb:')
        suburbLabel.grid(column=0, row=6)
        suburbEntry = tk.Entry(self, textvariable=self.suburb_var, width=round(self.winfo_screenwidth() * 0.03))
        suburbEntry.grid(column=1, row=6, sticky='W', padx=10, pady=5)

        states = ['VIC', 'ACT', 'NSW', 'QLD', 'SA', 'WA', 'TAS', 'NT']

        stateLabel = tk.Label(self, text='State:')
        stateLabel.grid(column=0, row=7)
        stateOption = tk.OptionMenu(self, self.state_var, *states)
        stateOption.grid(column=1, row=7, sticky='W', padx=10, pady=5)

        postcodeLabel = tk.Label(self, text='Postcode:')
        postcodeLabel.grid(column=0, row=8)
        postcodeEntry = tk.Entry(self, textvariable=self.postcode_var, width=round(self.winfo_screenwidth() * 0.03))
        postcodeEntry.grid(column=1, row=8, sticky='W', padx=10, pady=5)

        home_phoneLabel = tk.Label(self, text='Home Phone:')
        home_phoneLabel.grid(column=0, row=9)
        home_phoneEntry = tk.Entry(self, textvariable=self.home_phone_var, width=round(self.winfo_screenwidth() * 0.03))
        home_phoneEntry.grid(column=1, row=9, sticky='W', padx=10, pady=5)

        mobileLabel = tk.Label(self, text='Mobile Phone:')
        mobileLabel.grid(column=0, row=10)
        mobileEntry = tk.Entry(self, textvariable=self.mobile_phone_var, width=round(self.winfo_screenwidth() * 0.03))
        mobileEntry.grid(column=1, row=10, sticky='W', padx=10, pady=5)

        emailLabel = tk.Label(self, text='Email:')
        emailLabel.grid(column=0, row=11)
        emailEntry = tk.Entry(self, textvariable=self.email_var, width=round(self.winfo_screenwidth() * 0.03))
        emailEntry.grid(column=1, row=11, sticky='W', padx=10, pady=5)

        member_status = ['ACTIVE', 'INACTIVE']

        member_statusLabel = tk.Label(self, text='Member Status:')
        member_statusLabel.grid(column=0, row=12)
        member_status_Option = tk.OptionMenu(self, self.member_status_var, *member_status)
        member_status_Option.grid(column=1, row=12, sticky='W', padx=10, pady=5)

        cancel = tk.Button(self, text='Cancel', width=6, command=self.destroy,
                           activebackground=button_pressed_color)
        cancel.grid(column=0, row=13)
        cancel.bind('<Enter>', on_hover)
        cancel.bind('<Leave>', off_hover)

        submit = tk.Button(self, text='Submit', width=6, command=self.editMemberSubmit,
                           activebackground=button_pressed_color)
        submit.grid(column=1, row=13)
        submit.bind('<Enter>', on_hover)
        submit.bind('<Leave>', off_hover)

        self.attributes("-topmost", True)

    @activity_log
    def editMemberSubmit(self):
        fname = self.fname_var.get()
        lname = self.lname_var.get()
        partner = self.partner_var.get()
        address = self.address_var.get()
        suburb = self.suburb_var.get()
        state = self.state_var.get()
        postcode = self.postcode_var.get()
        home_phone = self.home_phone_var.get()
        mobile = self.mobile_phone_var.get()
        email = self.email_var.get()
        status = self.member_status_var.get()

        try:
            self.databaseConnection.insert(f'update members'
                                 f' set '
                                 f'member_fname = "{self.databaseConnection.sanitise(fname)}", '
                                 f'member_lname = "{self.databaseConnection.sanitise(lname)}", '
                                 f'partner_name = "{self.databaseConnection.sanitise(partner)}", '
                                 f'street_address = "{self.databaseConnection.sanitise(address)}", '
                                 f'suburb = "{self.databaseConnection.sanitise(suburb)}", '
                                 f'state = "{self.databaseConnection.sanitise(state)}", '
                                 f'postcode = "{self.databaseConnection.sanitise(postcode)}", '
                                 f'home_phone = "{self.databaseConnection.sanitise(home_phone)}", '
                                 f'mobile_phone = "{self.databaseConnection.sanitise(mobile)}", '
                                 f'email = "{self.databaseConnection.sanitise(email)}", '
                                 f'member_status = "{self.databaseConnection.sanitise(status)}" '
                                 f'where '
                                 f'member_no = {self.member_no}')
            self.databaseConnection.commit()
        except Exception as e:
            self.databaseConnection.rollback()
            print('Error Logged')
            logger.exception(e)
            messagebox.showwarning('Unknown Error',
                                   f'An unknown error occurred and has been logged. Report to developer.')
        else:
            self.member_page.update_window()
            self.destroy()


class InvoiceWindow(tk.Tk):
    @activity_log 
    def __init__(self, connection, main_menu, *args, **kwargs ):
        tk.Tk.__init__(self, *args, **kwargs)
        self.main_menu = main_menu
        self.report_callback_exception = log_unhandled_exception
        self.databaseConnection = connection
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure(2, weight=1)
        self.grid_columnconfigure(3, weight=1)
        self.grid_columnconfigure(4, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self.grid_rowconfigure(2, weight=1)
        self.grid_rowconfigure(3, weight=1)
        self.grid_rowconfigure(4, weight=1)
        self.grid_rowconfigure(5, weight=1)

        if self.databaseConnection != None:
            self.members = self.databaseConnection.query("select "
                                                     "concat_ws(' ',member_fname, member_lname) as n, "
                                                     "member_no from members where member_no != 1 "
                                                     "and"
                                                     " member_status = 'ACTIVE'")
        else:
            self.members = []
        self.items_price_var = tk.StringVar(self)
        self.item_qty_var = tk.StringVar(self)
        member_names = [m[0] for m in self.members]
        self.members = dict(self.members)
        if self.databaseConnection != None:
            self.items = self.databaseConnection.query("select item_description, item_code, item_value from item where hidden = 0;")
        else:
            self.items = []
        item_desc = [c[0] for c in self.items]
        self.items = [(c[0], (c[1], c[2])) for c in self.items]
        self.items = dict(self.items)
        x = self.winfo_screenwidth()
        y = self.winfo_screenheight()
        a = 0.4
        b = 0.5
        self.geometry(str(round(x * a)) +
                      'x'
                      + str(round(y * b)) +
                      '+'
                      + str(round((-a * x + x) / 2)) +
                      '+'
                      + str(round((-b * y + y) / 2)))
        self.title('Invoicing')

        self.member_var = tk.StringVar(self)
        self.member_var.set('Select Member')
        member_options = tk.OptionMenu(self, self.member_var, *member_names)
        member_options.config(width=round(self.winfo_screenwidth()*0.022))
        member_options.grid(row=1, column=0, sticky='NE', padx=self.winfo_screenwidth()*0.02)

        self.items_var = tk.StringVar(self)
        self.items_var.set('Select Invoice Item')
        items_options = tk.OptionMenu(self, self.items_var, *item_desc)
        items_options.config(width=round(self.winfo_screenwidth()*0.022))
        items_options.grid(row=2, column=0, sticky='EW', padx=self.winfo_screenwidth()*0.02)

        self.item_prices = []

        price_label = tk.Label(self, text='Price')
        price_label.grid(row=1, column=1)
        quant_label = tk.Label(self, text='Quantity')
        quant_label.grid(row=1, column=2)
        item_price_field = tk.Entry(self, textvariable=self.items_price_var)
        item_price_field.grid(row=2, column=1)
        item_qty_field = tk.Entry(self, textvariable=self.item_qty_var)
        item_qty_field.grid(row=2, column=2)

        item_add_button = tk.Button(self, text='+', command=self.add_item_command,
                                    activebackground = button_pressed_color)
        item_add_button.grid(row=2, column=3)
        item_add_button.bind('<Enter>', on_hover)
        item_add_button.bind('<Leave>', off_hover)


        item_remove_button = tk.Button(self, text='-', command=self.remove_item_command,
                                       activebackground = button_pressed_color)
        item_remove_button.grid(row=2, column=4)
        item_remove_button.bind('<Enter>', on_hover)
        item_remove_button.bind('<Leave>', off_hover)

        invoice_table(self, row=3, columnspan=5, pady=10)

        # invoice_font = font.Font(family='Courier', size=10)
        # self.invoice_table = Treeview(self, height=10)
        # self.invoice_table["columns"] = ('item_desc', 'unit_price', 'qty', 'subtotal')
        # self.invoice_table.column('#0', width=invoice_font.measure("Item Code"), stretch=tk.NO)
        # self.invoice_table.column('item_desc', width=invoice_font.measure("Annual Membership Renewal Fee"),stretch=tk.NO)
        # self.invoice_table.column('unit_price', width=invoice_font.measure("Unit Price"), stretch=tk.NO)
        # self.invoice_table.column('qty', width=invoice_font.measure(" QTY "), stretch=tk.NO)
        # self.invoice_table.column('subtotal', width=invoice_font.measure("Sub Total"), stretch=tk.NO)
        # self.invoice_table.heading('#0', text="Item Code")
        # self.invoice_table.heading('item_desc', text="Item Description")
        # self.invoice_table.heading('unit_price', text="Unit Price")
        # self.invoice_table.heading('qty', text="QTY")
        # self.invoice_table.heading('subtotal', text="Sub Total")
        # self.invoice_table.grid(columnspan=5, row=3, pady=10)

        self.invoice_total_var = tk.StringVar(self)
        self.invoice_total_var.set('$0.00')
        self.total = 0

        self.due_date = tk.StringVar(self)
        self.duedate = tk.Label(self, text='Due Date:')
        self.duedate.grid(row=4, column=0, sticky='E')

        self.duedate = DateEntry(self, date_pattern='dd/mm/yyyy')
        self.duedate.grid(row=4, column=1, sticky='w')

        self.total_price_entry = tk.Entry(self, textvariable=self.invoice_total_var, width=round(self.winfo_screenwidth() * 0.01))
        self.total_price_entry.grid(row=4, column=2, columnspan=2)

        total_label = tk.Label(self, text='Total:')
        total_label.grid(row=4, column=1, sticky='E')

        self.invoice_table.bind('<ButtonRelease-1>', self.select_item)

        submit_invoice_button = tk.Button(self, text='Submit Invoice', command=self.submit_invoice_command,
                                          activebackground = button_pressed_color)
        submit_invoice_button.grid(row=5, column=2, pady=self.winfo_screenheight()*0.005)
        submit_invoice_button.bind('<Enter>', on_hover)
        submit_invoice_button.bind('<Leave>', off_hover)


        cancel = tk.Button(self, text='Cancel', command=self.destroy, activebackground = button_pressed_color)
        cancel.grid(row=5, column=1)
        cancel.bind('<Enter>', on_hover)
        cancel.bind('<Leave>', off_hover)

        self.items_var.trace('w', self.set_price_qty)

    @activity_log
    def set_price_qty(self, *args):
        item = self.items_var.get()
        if item != 'Select Invoice Item':
            self.items_price_var.set(self.items[item][1])
            self.item_qty_var.set(1)

    @activity_log
    def select_item(self, a):
        self.row_values = self.invoice_table.item(self.invoice_table.selection())

    @activity_log
    def add_item_command(self):
        if self.member_var.get() != 'Select Member':
            item_selection = self.items_var.get()
            item_code = self.items[item_selection][0]
            item_value = self.items[item_selection][1]
            item_price = round(float(self.items_price_var.get()), 2)
            item_qty = int(self.item_qty_var.get())
            sub_total = item_price*item_qty
            row = (item_code, item_selection, item_price, item_qty, sub_total)
            self.item_prices.append(row)

            self.invoice_table.insert('', 'end', text=item_code, values=[item_selection, item_price, item_qty, sub_total])
            self.items_var.set('Select Invoice Item')
            self.items_price_var.set('')
            self.item_qty_var.set('')

            self.total = sum([i[-1] for i in self.item_prices])
            price_string = '$' + str(self.total)
            self.invoice_total_var.set(price_string)
            self.attributes("-topmost", True)
        else:
            messagebox.showwarning('Member not selected', 'A member must be selected first.', parent=self)

    @activity_log
    def remove_item_command(self):
        item_code = self.row_values['text']
        item_desc = self.row_values['values'][0]
        item_price = self.row_values['values'][1]
        item_qty = self.row_values['values'][2]
        sub_total = self.row_values['values'][3]
        row = (item_code, item_desc, float(item_price), item_qty, float(sub_total))
        self.item_prices.remove(row)
        self.invoice_table.delete(self.invoice_table.selection())

        self.total = sum([i[-1] for i in self.item_prices])
        total = str(self.total)
        if self.total == 0:
            total = '0.00'
        price_string = '$' + total
        self.invoice_total_var.set(price_string)

    @activity_log
    def submit_invoice_command(self):
        if self.databaseConnection != None:
                invoice_total = str(self.total)
                self.current_invoice_no = self.databaseConnection.query('select max(invoice_no)+1 from invoice;')

                q = invoice_total.split('.')
                if len(q[1]) == 1:
                    invoice_total = invoice_total + '0'

                self.current_invoice_no = self.current_invoice_no[0][0]

                member_no = self.members[self.member_var.get()]
                member_name = self.member_var.get()
                member_details = self.databaseConnection.query(
                    f'select street_address, suburb, postcode, state from members where member_no = {member_no}')
                member_details = member_details[0]

                pdf_name = f'{self.current_invoice_no}.pdf'
                due_date = self.duedate.get()

                invoice_copy = InvoiceGenerator(pdf_name, invoice_total, dueDate=due_date[:-4]+due_date[-2:])
                due_date = '"' + due_date + '"'

                bank_details = self.databaseConnection.query(f'select bank_name, bsb, account_no from bank_detail')
                bank_details = bank_details[0]

                item_codes = [r[0] for r in self.item_prices]
                item_descs = [r[1] for r in self.item_prices]
                item_prices = [r[2] for r in self.item_prices]

                for i in range(len(item_prices)):
                    q = str(item_prices[i]).split('.')
                    if len(q[1]) == 1:
                        s = str(item_prices[i]) + '0'
                        item_prices[i] = s

                item_qtys = [r[3] for r in self.item_prices]
                subtotals = [r[4] for r in self.item_prices]

                for i in range(len(subtotals)):
                    q = str(subtotals[i]).split('.')
                    if len(q[1]) == 1:
                        s = str(subtotals[i]) + '0'
                        subtotals[i] = s

                invoice_copy.memberName(member_name)
                invoice_copy.invoiceDate(datetime.today().date().strftime('%d-%m-%y'))
                invoice_copy.invoiceNo(self.current_invoice_no)
                invoice_copy.streetAdress(member_details[0])
                invoice_copy.cityStatePostCode(member_details[1], member_details[3], member_details[2])
                invoice_copy.BankDetails(bank_details[0], bank_details[1], bank_details[2])
                invoice_copy.invoice_line(item_codes, item_descs, item_prices, item_qtys, subtotals)
                try:
                    self.databaseConnection.insert(f'insert into invoice '
                                                   f'(invoice_no, '
                                                   f'invoice_date, '
                                                   f'invoice_duedate, '
                                                   f'invoice_total, '
                                                   f'member_no, '
                                                   f'invoice_sent) '
                                                   f'values '
                                                   f'({self.current_invoice_no}, '
                                                   f'now(), '
                                                   f'str_to_date({due_date},"%d/%m/%Y"), '
                                                   f'{invoice_total}, '
                                                   f'{member_no}, '
                                                   f'"No")')
                    for line in self.item_prices:
                        item_code = line[0]
                        item_price = line[2]
                        item_qty = line[3]

                        self.databaseConnection.insert(f'insert into invoice_line '
                                                   f'(invoice_no, item_code, item_qty, invoice_item_value) values '
                                                   f'({self.current_invoice_no}, {item_code}, {item_qty}, {item_price})')
                    self.databaseConnection.commit()
                except Exception as e:
                    self.databaseConnection.rollback()
                    print('Error Logged')
                    logger.exception(e)
                    messagebox.showwarning('Unknown Error',
                                           f'An unknown error occurred and has been logged. Report to developer.')
                else:
                    invoice_copy.save(f'{FOLDER}\\invoice_pdfs')
                    self.main_menu.update_tables()
                    self.destroy()
                    messagebox.showinfo('Invoice Created', 'Invoice successfully created', parent=self.main_menu)
        else:
            not_connected_message(self)


class NewMemberPage(tk.Tk):
    @activity_log
    def __init__(self, connection, member_page, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)
        height = self.winfo_screenheight()
        width = self.winfo_screenwidth()
        self.member_page = member_page
        self.report_callback_exception = log_unhandled_exception

        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self.grid_rowconfigure(2, weight=1)
        self.grid_rowconfigure(3, weight=1)
        self.grid_rowconfigure(4, weight=1)
        self.grid_rowconfigure(5, weight=1)
        self.grid_rowconfigure(6, weight=1)
        self.grid_rowconfigure(7, weight=1)
        self.grid_rowconfigure(8, weight=1)
        self.grid_rowconfigure(9, weight=1)
        self.grid_rowconfigure(10, weight=1)
        self.grid_rowconfigure(11, weight=1)
        self.grid_rowconfigure(12, weight=1)
        self.grid_rowconfigure(13, weight=1)

        self.geometry(str(round(width*0.3))+'x'+str(round(height*0.6)))
        self.dbconnection = connection
        # self.member_no_var = tk.StringVar(self)
        self.fname_var = tk.StringVar(self)
        self.lname_var = tk.StringVar(self)
        self.partner_var = tk.StringVar(self)
        self.address_var = tk.StringVar(self)
        self.suburb_var = tk.StringVar(self)
        self.state_var = tk.StringVar(self)
        self.postcode_var = tk.StringVar(self)
        self.home_phone_var = tk.StringVar(self)
        self.mobile_phone_var = tk.StringVar(self)
        self.email_var = tk.StringVar(self)
        self.member_status_var = tk.StringVar(self)


        self.title("Add New Member")
        titleLabel = tk.Label(self, text='Add New Member')
        titleLabel.grid(row=0, column=0)

        fnameLabel = tk.Label(self, text='First Name:')
        fnameLabel.grid(column=0, row=2)
        fnameEntry = tk.Entry(self, textvariable=self.fname_var, width=round(self.winfo_screenwidth() * 0.03))
        fnameEntry.grid(column=1, row=2, sticky='W', padx=10, pady=5)

        lnameLabel = tk.Label(self, text='Last Name:')
        lnameLabel.grid(column=0, row=3)
        lnameEntry = tk.Entry(self, textvariable=self.lname_var,width=round(self.winfo_screenwidth() * 0.03))
        lnameEntry.grid(column=1, row=3, sticky='W', padx=10, pady=5)

        partnerLabel = tk.Label(self, text='Partner Full name:')
        partnerLabel.grid(column=0, row=4)
        partnerEntry = tk.Entry(self, textvariable=self.partner_var, width=round(self.winfo_screenwidth() * 0.03))
        partnerEntry.grid(column=1, row=4, sticky='W', padx=10, pady=5)

        addressLabel = tk.Label(self, text='Street Address:')
        addressLabel.grid(column=0, row=5)
        addressEntry = tk.Entry(self, textvariable=self.address_var, width=round(self.winfo_screenwidth() * 0.03))
        addressEntry.grid(column=1, row=5, sticky='W', padx=10, pady=5)

        suburbLabel = tk.Label(self, text='Suburb:')
        suburbLabel.grid(column=0, row=6)
        suburbEntry = tk.Entry(self, textvariable=self.suburb_var, width=round(self.winfo_screenwidth() * 0.03))
        suburbEntry.grid(column=1, row=6, sticky='W', padx=10, pady=5)

        states = ['VIC', 'ACT', 'NSW', 'QLD', 'SA', 'WA', 'TAS', 'NT']

        stateLabel = tk.Label(self, text='State:')
        stateLabel.grid(column=0, row=7)
        stateOption = tk.OptionMenu(self, self.state_var, *states)
        stateOption.grid(column=1, row=7, sticky='W', padx=10, pady=5)

        postcodeLabel = tk.Label(self, text='Postcode:')
        postcodeLabel.grid(column=0, row=8)
        postcodeEntry = tk.Entry(self, textvariable=self.postcode_var, width=round(self.winfo_screenwidth() * 0.03))
        postcodeEntry.grid(column=1, row=8, sticky='W', padx=10, pady=5)

        home_phoneLabel = tk.Label(self, text='Home Phone:')
        home_phoneLabel.grid(column=0, row=9)
        home_phoneEntry = tk.Entry(self, textvariable=self.home_phone_var, width=round(self.winfo_screenwidth() * 0.03))
        home_phoneEntry.grid(column=1, row=9, sticky='W', padx=10, pady=5)

        mobileLabel = tk.Label(self, text='Mobile Phone:')
        mobileLabel.grid(column=0, row=10)
        mobileEntry = tk.Entry(self, textvariable=self.mobile_phone_var, width=round(self.winfo_screenwidth() * 0.03))
        mobileEntry.grid(column=1, row=10, sticky='W', padx=10, pady=5)

        emailLabel = tk.Label(self, text='Email:')
        emailLabel.grid(column=0, row=11)
        emailEntry = tk.Entry(self, textvariable=self.email_var, width=round(self.winfo_screenwidth() * 0.03))
        emailEntry.grid(column=1, row=11, sticky='W', padx=10, pady=5)

        member_status = ['ACTIVE', 'INACTIVE']

        member_statusLabel = tk.Label(self, text='Member Status:')
        member_statusLabel.grid(column=0, row=12)
        member_status_Option = tk.OptionMenu(self, self.member_status_var, *member_status)
        member_status_Option.grid(column=1, row=12, sticky='W', padx=10, pady=5)

        cancel = tk.Button(self, text='Cancel', width=6, command=self.destroy, activebackground = button_pressed_color)
        cancel.grid(column=0, row=13)
        cancel.bind('<Enter>', on_hover)
        cancel.bind('<Leave>', off_hover)


        submit = tk.Button(self, text='Submit', width=6, command=self.newMemberSubmit,
                           activebackground=button_pressed_color)
        submit.grid(column=1, row=13)
        submit.bind('<Enter>', on_hover)
        submit.bind('<Leave>', off_hover)

        self.attributes("-topmost", True)

    @activity_log
    def newMemberSubmit(self):
        if self.dbconnection != None:
            fname = self.fname_var.get()
            fname = fname.capitalize()
            lname = self.lname_var.get()
            lname = lname.capitalize()
            partner = self.partner_var.get()
            partner = partner.title()
            address = self.address_var.get()
            address = address.title()
            suburb = self.suburb_var.get()
            suburb = suburb.upper()
            state = self.state_var.get()
            postcode = self.postcode_var.get()
            home_phone = self.home_phone_var.get()

            mobile = self.mobile_phone_var.get()
            email = self.email_var.get()
            status = self.member_status_var.get()

            errors = []
            if not postcode.isnumeric() and len(postcode) > 0:
                errors.append('-Postcode\n')
                postcode_error = False
            else:
                postcode_error = True

            if not home_phone.isnumeric() and len(home_phone) > 0:
                errors.append('-Home Phone\n')
                home_phone_error = False
            else:
                home_phone_error = True

            if not mobile.isnumeric() and len(mobile) > 0:
                errors.append('-Mobile Phone\n')
                mobile_error = False
            else:
                mobile_error = True

            if status == None:
                messagebox.showwarning('Status Error', 'Status must be selected')
                status_error = False
            else:
                status_error = True

            if len(errors) > 0:
                error_str = ''.join(errors)

            if postcode_error and home_phone_error and mobile_error and status_error:
                try:
                    self.dbconnection.insert(f'insert'
                                             f' into members ( '
                                             f'member_fname, '
                                             f'member_lname, '
                                             f'partner_name, '
                                             f'street_address, '
                                             f'suburb, state, '
                                             f'postcode, '
                                             f'home_phone, '
                                             f'mobile_phone, '
                                             f'email, '
                                             f'member_status) '
                                             f'values ( '
                                             f'"{self.dbconnection.sanitise(fname)}", '
                                             f'"{self.dbconnection.sanitise(lname)}", '
                                             f'"{self.dbconnection.sanitise(partner)}", '
                                             f'"{self.dbconnection.sanitise(address)}", '
                                             f'"{self.dbconnection.sanitise(suburb)}", '
                                             f'"{self.dbconnection.sanitise(state)}", '
                                             f'"{self.dbconnection.sanitise(postcode)}", '
                                             f'"{self.dbconnection.sanitise(home_phone)}", '
                                             f'"{self.dbconnection.sanitise(mobile)}", '
                                             f'"{self.dbconnection.sanitise(email)}", '
                                             f'"{self.dbconnection.sanitise(status)}")')
                except Exception as e:
                    self.dbconnection.rollback()
                    print('Error Logged')
                    logger.exception(e)
                    messagebox.showwarning('Unknown Error',
                                           f'An unknown error occurred and has been logged. Report to developer.')
                else:
                    self.dbconnection.commit()
                    self.member_page.update_window()
                    self.destroy()
            else:
                messagebox.showwarning('Format Error', f'The following fields need to be numeric:\n\n{error_str}', parent=self)
        else:
            not_connected_message(self)


class HistoryWindow(tk.Tk):
    def __init__(self, receipt_no, connection, main_menu, type, email_address, email_password, email_host, email_port,*args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)
        state = 'on'
        self.report_callback_exception = log_unhandled_exception
        self.receipt_no = int(receipt_no)
        self.main_menu = main_menu
        self.databaseConnection = connection
        self.receipt_type = type
        self.email_address = email_address
        self.email_host = email_host
        self.email_port = email_port
        self.email_password = email_password
        # Invoice history
        if self.receipt_no >= 40000 and self.receipt_no < 50000:
            self.grid_columnconfigure(0, weight=1)
            self.grid_columnconfigure(1, weight=1)
            self.grid_columnconfigure(2, weight=1)
            self.grid_rowconfigure(0, weight=1)
            self.grid_rowconfigure(1, weight=1)
            self.grid_rowconfigure(2, weight=1)
            self.grid_rowconfigure(3, weight=1)
            self.grid_rowconfigure(4, weight=1)
            self.grid_rowconfigure(5, weight=1)
            self.grid_rowconfigure(6, weight=1)
            self.grid_rowconfigure(7, weight=1)
            self.grid_rowconfigure(8, weight=1)
            self.grid_rowconfigure(9, weight=1)
            x = self.winfo_screenwidth()
            y = self.winfo_screenheight()
            a = 0.45
            b = 0.65
            self.title('Invoice')
            self.geometry(str(round(x * a)) +
                          'x'
                          + str(round(y * b)) +
                          '+'
                          + str(round((-a * x + x) / 2)) +
                          '+'
                          + str(round((-b * y + y) / 2)))
            try:
                self.invoice_data = self.databaseConnection.query(f'select '
                                                                  f'invoice_receipt_no, ' 
                                                                  f'invoice.invoice_no, '
                                                                  f'date_format(payment_datetime, "%d-%m-%Y %T"), '
                                                                  f'date_format(invoice_date, "%d-%m-%Y %T"), '
                                                                  f'concat_ws(" ", member_fname, member_lname), '
                                                                  f'cash_amount, '
                                                                  f'transfer_amount, '
                                                                  f'invoice_total,'
                                                                  f'email, '
                                                                  f'receipt_sent, '
                                                                  f'street_address, '
                                                                  f'state, '
                                                                  f'postcode,'
                                                                  f'suburb '
                                                                  f'from '
                                                                  f'invoice_receipt '
                                                                  f'join '
                                                                  f'invoice '
                                                                  f'on '
                                                                  f'invoice.invoice_no = invoice_receipt.invoice_no '
                                                                  f'join '
                                                                  f'members '
                                                                  f'on '
                                                                  f'members.member_no = invoice.member_no '
                                                                  f'where '
                                                                  f'invoice_receipt_no = {receipt_no};')
            except Exception as e:
                self.destroy()
                state = 'off'
                logger.exception(e)
            if state == 'on':
                self.invoice_data = self.invoice_data[0]
                try:
                    self.invoice_line = self.databaseConnection.query(f'select '
                                                                  f'item.item_code, '
                                                                  f'item_description, '
                                                                  f'invoice_item_value, '
                                                                  f'item_qty, '
                                                                  f'item_qty*invoice_item_value '
                                                                  f'from '
                                                                  f'invoice_line '
                                                                  f'join '
                                                                  f'item '
                                                                  f'on '
                                                                  f'invoice_line.item_code = item.item_code '
                                                                  f'where invoice_no = {self.invoice_data[1]} ;')
                except:
                    state = 'off'
                    self.destroy()
                if state == 'on':
                    self.receipt_data = {'receipt_no': self.invoice_data[0], 'invoice_no': self.invoice_data[1],
                                         'payment_datetime': self.invoice_data[2],
                                         'issue_datetime': self.invoice_data[3], 'fullname': self.invoice_data[4],
                                         'cash_amount': self.invoice_data[5], 'transfer_amount': self.invoice_data[6],
                                         'invoice_total': self.invoice_data[7], 'member_email': self.invoice_data[8],
                                         'email_sent': self.invoice_data[9], 'street_address': self.invoice_data[10],
                                         'state': self.invoice_data[11], 'postcode': self.invoice_data[12],
                                         'suburb': self.invoice_data[13]}

                    self.invoice_heading = tk.Label(self, text=f'Invoice Receipt: {self.receipt_no}', font='Courier 30 bold')
                    self.invoice_heading.grid(row=0, columnspan=3, sticky='w')

                    self.member_name_label = tk.Label(self, text=f'Member Name: {self.receipt_data["fullname"]}', font='Courier 15 bold')
                    self.member_name_label.grid(row=1, columnspan=3, sticky='nw', padx=10)

                    self.issue_date_label = tk.Label(self, text=f'Issue Date: {self.receipt_data["issue_datetime"]}', font='Courier 15 bold')
                    self.issue_date_label.grid(row=2, columnspan=3, sticky='nw', padx=10)

                    self.payment_date_label = tk.Label(self, text=f'Payment Date: {self.receipt_data["payment_datetime"]}', font='Courier 15 bold')
                    self.payment_date_label.grid(row=3, columnspan=3, sticky='nw', padx=10)

                    self.total_label = tk.Label(self, text=f'Cash Amount: ${self.receipt_data["cash_amount"]} ', font='Courier 15 bold')
                    self.total_label.grid(row=4, columnspan=2, sticky='w', padx=10)

                    self.total_label = tk.Label(self, text=f'Transfer Amount: ${self.receipt_data["transfer_amount"]} ', font='Courier 15 bold')
                    self.total_label.grid(row=5, columnspan=2, sticky='w', padx=10)

                    self.total_label = tk.Label(self, text=f'Total: ${self.receipt_data["invoice_total"]} ', font='Courier 15 bold')
                    self.total_label.grid(row=6, columnspan=2, sticky='w', padx=10)

                    if self.receipt_data["email_sent"] == 'No':
                        self.send_invoice_button = tk.Button(self, text='Send Receipt', command=self.send_receipt,
                                                             activebackground = button_pressed_color)
                        self.send_invoice_button.bind('<Enter>', on_hover)
                        self.send_invoice_button.bind('<Leave>', off_hover)
                        button_font = font.Font(family='Courier', size=15, weight='bold')
                        self.send_invoice_button['font'] = button_font
                        self.send_invoice_button.grid(row=6, column=2)
                    else:
                        self.send_invoice_button = tk.Button(self, text='Resend Receipt', command=self.send_receipt,
                                                             activebackground = button_pressed_color)
                        button_font = font.Font(family='Courier', size=15, weight='bold')
                        self.send_invoice_button['font'] = button_font
                        self.send_invoice_button.grid(row=6, column=2)
                        self.send_invoice_button.bind('<Enter>', on_hover)
                        self.send_invoice_button.bind('<Leave>', off_hover)


                    # self.receipt_table = Treeview(self)
                    # self.receipt_table["columns"] = ('item_desc', 'unit_price', 'qty', 'subtotal')
                    # self.receipt_table.column('#0', width=round(self.winfo_screenwidth() * 0.05), minwidth=50, stretch=tk.NO)
                    # self.receipt_table.column('item_desc', width=round(self.winfo_screenwidth() * 0.12), minwidth=50,
                    #                           stretch=tk.NO)
                    # self.receipt_table.column('unit_price', width=round(self.winfo_screenwidth() * 0.07), minwidth=50,
                    #                           stretch=tk.NO)
                    # self.receipt_table.column('qty', width=round(self.winfo_screenwidth() * 0.06), minwidth=50, stretch=tk.NO)
                    # self.receipt_table.column('subtotal', width=round(self.winfo_screenwidth() * 0.06), minwidth=50,
                    #                           stretch=tk.NO)
                    #
                    # self.receipt_table.heading('#0', text="Item Code")
                    # self.receipt_table.heading('item_desc', text="Item Description")
                    # self.receipt_table.heading('unit_price', text="Unit Price")
                    # self.receipt_table.heading('qty', text="QTY")
                    # self.receipt_table.heading('subtotal', text="Sub Total")
                    #
                    # self.receipt_table.grid(row=7, columnspan=3, pady=(0,10))
                    invoice_table(self, row=7, columnspan=3, pady=(0,10))

                    for row in self.invoice_line:
                        self.invoice_table.insert('', 'end', text=row[0], values=row[1:])

                    self.refund_button = tk.Button(self, text='Refund', command=self.refund_invoice,
                                                   activebackground = button_pressed_color)
                    self.refund_button['font'] = button_font
                    self.refund_button.grid(row=8, column=0, pady=10)
                    self.refund_button.bind('<Enter>', on_hover)
                    self.refund_button.bind('<Leave>', off_hover)
                    activebackground = button_pressed_color

                    self.delete_button = tk.Button(self, text='Delete', command=self.delete_receipt,
                                                   activebackground = button_pressed_color)
                    self.delete_button['font'] = button_font
                    self.delete_button.grid(row=8, column=1, pady=10)
                    self.delete_button.bind('<Enter>', on_hover)
                    self.delete_button.bind('<Leave>', off_hover)


                    self.view_receipt_button = tk.Button(self, text='View Receipt', command=self.view_receipt)
                    self.view_receipt_button['font'] = button_font
                    self.view_receipt_button.grid(row=8, column=2, pady=10)
                    self.view_receipt_button.bind('<Enter>', on_hover)
                    self.view_receipt_button.bind('<Leave>', off_hover)


                    self.attributes("-topmost", True)  # #
                else:
                    messagebox.showerror('Error', 'Error occurred. Report to developer.')
            else:
                messagebox.showerror('Error', 'Error occurred. Report to developer.')
        # income history
        elif self.receipt_no > 30000 and self.receipt_no < 40000:
            self.grid_columnconfigure(0, weight=1)
            self.grid_columnconfigure(1, weight=1)
            self.grid_columnconfigure(2, weight=1)
            self.grid_rowconfigure(0, weight=1)
            self.grid_rowconfigure(1, weight=1)
            self.grid_rowconfigure(2, weight=1)
            self.grid_rowconfigure(3, weight=1)
            self.grid_rowconfigure(4, weight=1)
            self.grid_rowconfigure(5, weight=1)
            self.grid_rowconfigure(6, weight=1)
            x = self.winfo_screenwidth()
            y = self.winfo_screenheight()
            a = 0.45
            b = 0.35
            self.title('Income')
            self.geometry(str(round(x * a)) +
                          'x'
                          + str(round(y * b)) +
                          '+'
                          + str(round((-a * x + x) / 2)) +
                          '+'
                          + str(round((-b * y + y) / 2)))
            income_data = self.databaseConnection.query('select '
                                                         'cash_amount, '
                                                         'transfer_amount, '
                                                         'payment_datetime, '
                                                         'income_notes, '
                                                         'income_description '
                                                         'from '
                                                         'income_receipt '
                                                         'join income '
                                                         'on '
                                                         'income_receipt.income_id = income.income_id '
                                                         f'where income_receipt_no = {receipt_no}')

            income_data = income_data[0]

            self.invoice_heading = tk.Label(self, text=f'Club Income: {self.receipt_no}', font='Courier 30 bold')
            self.invoice_heading.grid(row=0, columnspan=3, sticky='w')

            self.income_date_label = tk.Label(self, text=f'Expense Date: {income_data[2]}',
                                              font='Courier 15 bold')
            self.income_date_label.grid(row=1, columnspan=3, sticky='nw')

            self.income_decription_label = tk.Label(self, text=f'Category: {income_data[4]}',
                                                    font='Courier 15 bold')
            self.income_decription_label.grid(row=2, columnspan=3, sticky='nw')

            self.cash_label = tk.Label(self, text=f'Cash Amount: ${income_data[0]}', font='Courier 15 bold')
            self.cash_label.grid(row=3, columnspan=3, sticky='nw')

            self.transfer_label = tk.Label(self, text=f'Transfer Amount: ${income_data[1]} ',
                                           font='Courier 15 bold')
            self.transfer_label.grid(row=4, columnspan=3, sticky='nw')

            self.notes_label = tk.Label(self, text='Notes', font='Courier 15 bold')
            self.notes_label.grid(row=5, column=0, sticky='sw')
            self.notes = tk.Text(self, width=60, height=7)
            self.notes.insert('1.0', f'{income_data[3]}')
            self.notes.grid(row=6, columnspan=3, pady=(0, 20))

            self.delete_button = tk.Button(self, text='Delete', command=self.delete_income,
                                           activebackground = button_pressed_color)
            self.delete_button.grid(row=7, columnspan=3)
            self.delete_button.bind('<Enter>', on_hover)
            self.delete_button.bind('<Leave>', off_hover)

            self.attributes("-topmost", True)
        # cash out or deposit
        elif self.receipt_no > 50000 and self.receipt_no < 70000:
            x = self.winfo_screenwidth()
            y = self.winfo_screenheight()
            a = 0.45
            b = 0.35
            self.geometry(str(round(x * a)) +
                          'x'
                          + str(round(y * b)) +
                          '+'
                          + str(round((-a * x + x) / 2)) +
                          '+'
                          + str(round((-b * y + y) / 2)))
            self.grid_columnconfigure(0, weight=1)
            self.grid_columnconfigure(1, weight=1)
            self.grid_columnconfigure(2, weight=1)
            self.grid_rowconfigure(0, weight=1)
            self.grid_rowconfigure(1, weight=1)
            self.grid_rowconfigure(2, weight=1)
            self.grid_rowconfigure(3, weight=1)
            self.grid_rowconfigure(4, weight=1)
            transfer_data = self.databaseConnection.query('select '
                                                          'cash_amount, '
                                                          'transfer_amount, '
                                                          'payment_datetime '
                                                          'from '
                                                          'transfer '
                                                          f'where transfer_no = {receipt_no}')
            transfer_data = transfer_data[0]
            if type == 'Cash out':
                self.title('Cash out')
                self.invoice_heading = tk.Label(self, text=f'Cash Out: {self.receipt_no}', font='Courier 30 bold')
                self.invoice_heading.grid(row=0, columnspan=3, sticky='w')
            elif type == 'Deposit':
                self.title('Deposit')
                self.invoice_heading = tk.Label(self, text=f'Bank Deposit: {self.receipt_no}', font='Courier 30 bold')
                self.invoice_heading.grid(row=0, columnspan=3, sticky='w')

            self.expense_date_label = tk.Label(self, text=f'Transfer Date: {transfer_data[2]}',
                                               font='Courier 15 bold')
            self.expense_date_label.grid(row=1, columnspan=3, sticky='nw')

            self.cash_label = tk.Label(self, text=f'Cash Amount: ${transfer_data[0]}', font='Courier 15 bold')
            self.cash_label.grid(row=2, columnspan=3, sticky='nw')

            self.transfer_label = tk.Label(self, text=f'Transfer Amount: ${transfer_data[1]} ',
                                           font='Courier 15 bold')
            self.transfer_label.grid(row=3, columnspan=3, sticky='nw')

            self.attributes("-topmost", True)
            button_font = font.Font(family='Courier', size=15, weight='bold')
            self.delete_transfer_button = tk.Button(self, text='Delete', command=self.delete_transfer_record,
                                                    activebackground = button_pressed_color)
            self.delete_transfer_button['font'] = button_font
            self.delete_transfer_button.grid(row=4, columnspan=3)
            self.delete_transfer_button.bind('<Enter>', on_hover)
            self.delete_transfer_button.bind('<Leave>', off_hover)
        # expense history
        elif self.receipt_no > 70000 and self.receipt_no < 90000:
            x = self.winfo_screenwidth()
            y = self.winfo_screenheight()
            a = 0.45
            b = 0.35
            self.geometry(str(round(x * a)) +
                          'x'
                          + str(round(y * b)) +
                          '+'
                          + str(round((-a * x + x) / 2)) +
                          '+'
                          + str(round((-b * y + y) / 2)))
            self.grid_columnconfigure(0, weight=1)
            self.grid_columnconfigure(1, weight=1)
            self.grid_columnconfigure(2, weight=1)
            self.grid_rowconfigure(0, weight=1)
            self.grid_rowconfigure(1, weight=1)
            self.grid_rowconfigure(2, weight=1)
            self.grid_rowconfigure(3, weight=1)
            self.grid_rowconfigure(4, weight=1)
            self.grid_rowconfigure(5, weight=1)
            self.grid_rowconfigure(6, weight=1)
            self.grid_rowconfigure(7, weight=1)
            expense_data = self.databaseConnection.query('select '
                                                         'cash_amount, '
                                                         'transfer_amount, '
                                                         'payment_datetime, '
                                                         'expense_notes, '
                                                         'expense_description '
                                                         'from '
                                                         'expense_receipt '
                                                         'join expense '
                                                         'on '
                                                         'expense_receipt.expense_id = expense.expense_id '
                                                         f'where expense_receipt_no = {receipt_no}')
            expense_data = expense_data[0]
            self.title('Expense')
            self.invoice_heading = tk.Label(self, text=f'Club Expense: {self.receipt_no}', font='Courier 30 bold')
            self.invoice_heading.grid(row=0, columnspan=3, sticky='w')

            self.expense_date_label = tk.Label(self, text=f'Expense Date: {expense_data[2]}', font='Courier 15 bold')
            self.expense_date_label.grid(row=1, columnspan=3, sticky='nw')

            self.expense_decription_label = tk.Label(self, text=f'Category: {expense_data[4]}', font='Courier 15 bold')
            self.expense_decription_label.grid(row=2, columnspan=3, sticky='nw')

            self.cash_label = tk.Label(self, text=f'Cash Amount: ${expense_data[0]}', font='Courier 15 bold')
            self.cash_label.grid(row=3, columnspan=3, sticky='nw')

            self.transfer_label = tk.Label(self, text=f'Transfer Amount: ${expense_data[1]} ', font='Courier 15 bold')
            self.transfer_label.grid(row=4, columnspan=3, sticky='nw')


            self.notes_label = tk.Label(self, text='Notes', font='Courier 15 bold')
            self.notes_label.grid(row=5, column=0, sticky='sw')
            self.notes = tk.Text(self, width=60, height=7)
            self.notes.insert('1.0', f'{expense_data[3]}')
            self.notes.grid(row=6, columnspan=3, pady=(0,20))

            self.delete_button = tk.Button(self, text='Delete', command=self.delete_expense,
                                           activebackground = button_pressed_color)
            self.delete_button.grid(row=7, columnspan=3)
            self.delete_button.bind('<Enter>', on_hover)
            self.delete_button.bind('<Leave>', off_hover)

            self.attributes("-topmost", True)
        # refunds history
        elif self.receipt_no > 90000:
            self.grid_columnconfigure(0, weight=1)
            self.grid_columnconfigure(1, weight=1)
            self.grid_columnconfigure(2, weight=1)
            self.grid_rowconfigure(0, weight=1)
            self.grid_rowconfigure(1, weight=1)
            self.grid_rowconfigure(2, weight=1)
            self.grid_rowconfigure(3, weight=1)
            self.grid_rowconfigure(4, weight=1)
            self.grid_rowconfigure(5, weight=1)
            self.grid_rowconfigure(6, weight=1)
            self.grid_rowconfigure(7, weight=1)
            self.grid_rowconfigure(8, weight=1)
            self.grid_rowconfigure(9, weight=1)
            x = self.winfo_screenwidth()
            y = self.winfo_screenheight()
            a = 0.45
            b = 0.65
            self.geometry(str(round(x * a)) +
                          'x'
                          + str(round(y * b)) +
                          '+'
                          + str(round((-a * x + x) / 2)) +
                          '+'
                          + str(round((-b * y + y) / 2)))
            self.title('Refund')
            try:
                self.invoice_data = self.databaseConnection.query(f'select '
                                                                  f'refunds.invoice_receipt_no, '
                                                                  f'invoice_receipt.invoice_no, '
                                                                  f'refund_no, '
                                                                  f'date_format(invoice_receipt.payment_datetime, "%d-%m-%Y %T"), '
                                                                  f'date_format(invoice_date, "%d-%m-%Y %T"), '
                                                                  f'date_format(refunds.payment_datetime, "%d-%m-%Y %T"), '
                                                                  f'concat_ws(" ", member_fname, member_lname), '
                                                                  f'invoice_receipt.cash_amount, '
                                                                  f'invoice_receipt.transfer_amount, '
                                                                  f'invoice_total,'
                                                                  f'email, '
                                                                  f'refund_sent, '
                                                                  f'refunds.cash_amount, '
                                                                  f'refunds.transfer_amount '
                                                                  f'from '
                                                                  f'refunds '
                                                                  f'join invoice_receipt on refunds.invoice_receipt_no = invoice_receipt.invoice_receipt_no '
                                                                  f'join invoice on invoice_receipt.invoice_no = invoice.invoice_no '
                                                                  f'join members on invoice.member_no = members.member_no '
                                                                  f'where '
                                                                  f'refunds.refund_no = {self.receipt_no};')
            except Exception as e:
                self.destroy()
                state = 'off'
                logger.exception(e)
            if state == 'on':
                self.invoice_data = self.invoice_data[0]
                try:
                    self.invoice_line = self.databaseConnection.query(f'select '
                                                                      f'item.item_code, '
                                                                      f'item_description, '
                                                                      f'invoice_item_value, '
                                                                      f'item_qty, '
                                                                      f'item_qty*invoice_item_value '
                                                                      f'from '
                                                                      f'invoice_line '
                                                                      f'join '
                                                                      f'item '
                                                                      f'on '
                                                                      f'invoice_line.item_code = item.item_code '
                                                                      f'where invoice_no = {self.invoice_data[1]} ;')
                except:
                    state = 'off'
                    self.destroy()
                if state == 'on':
                    self.receipt_data = {'receipt_no': self.invoice_data[0], 'invoice_no': self.invoice_data[1],
                                         'payment_datetime': self.invoice_data[3], 'refund_no' : self.invoice_data[2],
                                         'issue_datetime': self.invoice_data[4], 'fullname': self.invoice_data[6],
                                         'cash_amount': self.invoice_data[7], 'transfer_amount': self.invoice_data[8],
                                         'invoice_total': self.invoice_data[9], 'member_email': self.invoice_data[10],
                                         'email_sent': self.invoice_data[11], 'refund_datetime': self.invoice_data[5],
                                         'refund_cash': self.invoice_data[12], 'refund_transfer':self.invoice_data[13]}

                    self.invoice_heading = tk.Label(self, text=f'Refund No: {self.receipt_no}',
                                                    font='Courier 30 bold')
                    self.invoice_heading.grid(row=0, columnspan=3, sticky='w')

                    self.member_name_label = tk.Label(self, text=f'Member Name: {self.receipt_data["fullname"]}',
                                                      font='Courier 15 bold')
                    self.member_name_label.grid(row=1, columnspan=3, sticky='nw', padx=10)

                    self.member_name_label = tk.Label(self, text=f'Receipt No: {self.receipt_data["receipt_no"]}',
                                                      font='Courier 15 bold')
                    self.member_name_label.grid(row=2, columnspan=3, sticky='nw', padx=10)

                    self.issue_date_label = tk.Label(self, text=f'Refund Date: {self.receipt_data["refund_datetime"]}',
                                                     font='Courier 15 bold')
                    self.issue_date_label.grid(row=3, columnspan=3, sticky='nw', padx=10)

                    self.total_label = tk.Label(self, text=f'Refund Cash Amount: ${self.receipt_data["refund_cash"]} ',
                                                font='Courier 15 bold')
                    self.total_label.grid(row=4, columnspan=2, sticky='w', padx=10)

                    self.total_label = tk.Label(self, text=f'Refund Transfer Amount: ${self.receipt_data["refund_transfer"]} ',
                                                font='Courier 15 bold')
                    self.total_label.grid(row=5, columnspan=2, sticky='w', padx=10)

                    self.total_label = tk.Label(self, text=f'Total: ${self.receipt_data["invoice_total"]} ',
                                                font='Courier 15 bold')
                    self.total_label.grid(row=6, columnspan=2, sticky='w', padx=10)

                    if self.receipt_data["email_sent"] == 'no':
                        self.send_invoice_button = tk.Button(self, text='Send Refund', command=self.send_refund,
                                                             activebackground = button_pressed_color)
                        button_font = font.Font(family='Courier', size=15, weight='bold')
                        self.send_invoice_button['font'] = button_font
                        self.send_invoice_button.grid(row=6, column=2)
                        self.send_invoice_button.bind('<Enter>', on_hover)
                        self.send_invoice_button.bind('<Leave>', off_hover)
                    else:
                        self.send_invoice_button = tk.Button(self, text='Resend Refund', command=self.send_refund,
                                                             activebackground = button_pressed_color)
                        button_font = font.Font(family='Courier', size=15, weight='bold')
                        self.send_invoice_button['font'] = button_font
                        self.send_invoice_button.grid(row=6, column=2)
                        self.send_invoice_button.bind('<Enter>', on_hover)
                        self.send_invoice_button.bind('<Leave>', off_hover)

                    # self.receipt_table = Treeview(self)
                    # self.receipt_table["columns"] = ('item_desc', 'unit_price', 'qty', 'subtotal')
                    # self.receipt_table.column('#0', width=round(self.winfo_screenwidth() * 0.05), minwidth=50,
                    #                           stretch=tk.NO)
                    # self.receipt_table.column('item_desc', width=round(self.winfo_screenwidth() * 0.12), minwidth=50,
                    #                           stretch=tk.NO)
                    # self.receipt_table.column('unit_price', width=round(self.winfo_screenwidth() * 0.07), minwidth=50,
                    #                           stretch=tk.NO)
                    # self.receipt_table.column('qty', width=round(self.winfo_screenwidth() * 0.06), minwidth=50,
                    #                           stretch=tk.NO)
                    # self.receipt_table.column('subtotal', width=round(self.winfo_screenwidth() * 0.06), minwidth=50,
                    #                           stretch=tk.NO)
                    #
                    # self.receipt_table.heading('#0', text="Item Code")
                    # self.receipt_table.heading('item_desc', text="Item Description")
                    # self.receipt_table.heading('unit_price', text="Unit Price")
                    # self.receipt_table.heading('qty', text="QTY")
                    # self.receipt_table.heading('subtotal', text="Sub Total")
                    #
                    # self.receipt_table.grid(row=7, columnspan=3, pady=(0, 10))

                    invoice_table(self, row=7, columnspan=3, pady=(0,10))

                    for row in self.invoice_line:
                        self.invoice_table.insert('', 'end', text=row[0], values=row[1:])

                    self.attributes("-topmost", True)  # #
                else:
                    messagebox.showerror('Error', 'Error occurred. Report to developer.')
            else:
                messagebox.showerror('Error', 'Error occurred. Report to developer.')

    @activity_log
    def delete_transfer_record(self):
        delete_yesno = messagebox.askyesno('Delete Transfer', 'Are you sure you want to delete this transfer record?', parent=self)
        if delete_yesno:
            self.databaseConnection.insert(f'delete from transfer where transfer_no = {self.receipt_no}')
            self.databaseConnection.commit()
            self.main_menu.update_tables()
            self.destroy()
            messagebox.showinfo('Delete Transfer', 'Transfer record deleted successfully', parent=self.main_menu)

    @activity_log
    def refund_invoice(self):
        refunded = self.databaseConnection.query(f'select invoice_receipt_no from refunds where invoice_receipt_no = {self.receipt_no}')
        if len(refunded) != 0:
            refund_data = self.databaseConnection.query(f'select '
                                                        f'refund_no, '
                                                        f'cash_amount+transfer_amount, '
                                                        f'payment_datetime '
                                                        f'from '
                                                        f'refunds where invoice_receipt_no = {self.receipt_no}')
            refund_data = refund_data[0]
            messagebox.showwarning('Already Refunded', f'Invoice no {self.receipt_data["invoice_no"]} (${refund_data[1]}) '
                                                       f'has already been refunded at {refund_data[2]}.', parent=self)
        else:
            RefundCashOrTransfer(self.receipt_no, self.databaseConnection, self.main_menu, self.receipt_data, self.invoice_line, self)

    @activity_log
    def delete_receipt(self):
        refunded = self.databaseConnection.query(
            f'select refund_no, invoice_receipt_no from refunds where invoice_receipt_no = {self.receipt_no}')
        if len(refunded) != 0:
            messagebox.showwarning("Receipt already refunded", f"Cannot delete receipt no {self.receipt_no} "
                                                               f"because it has already been refunded. "
                                                               f"(Refund no: {refunded[0][0]}) ", parent=self)
        else:
            delete_yesno = messagebox.askyesno('Delete Receipt', 'Are you sure you want to delete this receipt?', parent=self)
            if delete_yesno:
                self.databaseConnection.insert(f'delete from invoice_receipt where invoice_receipt_no = {self.receipt_no}')
                self.databaseConnection.commit()
                self.main_menu.update_tables()
                self.destroy()
                messagebox.showinfo('Receipt Deleted', 'Invoice receipt record deleted.', parent=self.main_menu)

    @activity_log
    def delete_income(self):
        delete_yesno = messagebox.askyesno('Delete Income Entry', 'Are you sure you want to delete this entry?', parent=self)
        if delete_yesno:
            self.databaseConnection.insert(f'delete from income_receipt where income_receipt_no = {self.receipt_no}')
            self.databaseConnection.commit()
            self.main_menu.update_tables()
            self.destroy()
            messagebox.showinfo('Entry Deleted', 'Income entry record deleted.', parent=self.main_menu)

    @activity_log
    def delete_expense(self):
        delete_yesno = messagebox.askyesno('Delete Expense Entry', 'Are you sure you want to delete this entry?', parent=self)
        if delete_yesno:
            self.databaseConnection.insert(f'delete from expense_receipt where expense_receipt_no = {self.receipt_no}')
            self.databaseConnection.commit()
            self.main_menu.update_tables()
            self.destroy()
            messagebox.showinfo('Entry Deleted', 'Expense entry record deleted.', parent=self.main_menu)

    @activity_log
    def view_receipt(self):
        self.attributes("-topmost", False)  # #
        startfile(f'{FOLDER}\\receipt_pdfs\\R{self.receipt_no}.pdf')

    @activity_log
    def send_receipt(self):
        email_yesno = messagebox.askyesno('Email Receipt',
                                          f'Would you like to send out this invoice to \n{self.receipt_data["member_email"]}?',
                                          parent=self)
        if email_yesno:
            print(f'sending receipt to {self.receipt_data["email_sent"]}')
            sender = Emailer('Greater Western 4x4 Club Receipt',  # subject
                    f'To {self.receipt_data["fullname"]},\n\n'  # body
                    f' See attached your Greater Western 4x4 receipt.\n\n',
                    self.email_address,  # sender
                    self.receipt_data["member_email"],  # receiver
                    self.email_password, self.email_host, self.email_port,  # email password
                    attachment_path=f'{FOLDER}/receipt_pdfs/R{self.receipt_no}.pdf', filename=f'R{self.receipt_no}.pdf')  # receipt attachment
            status = sender.get_status()
            if status == '1':
                self.databaseConnection.insert(
                    f'update invoice_receipt set receipt_sent = "Yes" where invoice_receipt_no = {self.receipt_no}')
                self.databaseConnection.commit()
                self.main_menu.update_tables()
                messagebox.showinfo('Emailed Receipt', 'Receipt sent successfully.', parent=self)
                self.destroy()
                HistoryWindow(self.receipt_no, self.databaseConnection, self.main_menu, self.receipt_type, self.email_address, self.email_password, self.email_host, self.email_port)
            else:
                logger.exception(sender.return_error())
                messagebox.showerror('Email failure', 'Email failed to send.', parent=self)
        else:
            print('Cancelled')

    @activity_log
    def send_refund(self):
        email_yesno = messagebox.askyesno('Email Refund',
                                          f'Would you like to send out this refund receipt to \n{self.receipt_data["member_email"]}?',
                                          parent=self)
        if email_yesno:
            print(f'sending refund receipt to {self.receipt_data["member_email"]}')
            sender = Emailer('Greater Western 4x4 Club Refund Receipt ',  # subject
                             f'To {self.receipt_data["fullname"]},\n\n'  # body
                             f' See attached your Greater Western 4x4 refund receipt.\n\n',
                             self.email_address,  # sender
                             self.receipt_data["member_email"],  # receiver
                             self.email_password, self.email_host, self.email_port,  # email password
                             attachment_path=f'{FOLDER}/refund_pdfs/{self.receipt_no}.pdf',
                             filename=f'{self.receipt_no}.pdf')  # receipt attachment
            status = sender.get_status()
            if status == '1':
                self.databaseConnection.insert(
                    f'update refunds set refund_sent = "yes" where refund_no = {self.receipt_no}')
                self.databaseConnection.commit()
                self.main_menu.update_tables()
                messagebox.showinfo('Emailed Receipt', 'Receipt sent successfully.', parent=self)
                self.destroy()
                HistoryWindow(self.receipt_no, self.databaseConnection, self.main_menu, self.receipt_type,
                              self.email_address, self.email_password, self.email_host, self.email_port)
            else:
                logger.exception(sender.return_error())
                messagebox.showerror('Email failure', 'Email failed to send', parent=self)
        else:
            print('Cancelled')


class RefundCashOrTransfer(tk.Tk):
    @activity_log
    def __init__(self, receipt_no, connection, main_menu, invoice_data, invoice_line, history_window,  *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)
        self.attributes("-topmost", True)
        self.invoice_no = receipt_no
        self.report_callback_exception = log_unhandled_exception
        self.main_menu = main_menu
        self.databaseConnection = connection
        self.receipt_data = invoice_data
        self.receipt_line = invoice_line
        self.history_window = history_window
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure(2, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self.grid_rowconfigure(2, weight=1)
        x = self.winfo_screenwidth()
        y = self.winfo_screenheight()
        a = 0.3
        b = 0.2
        self.geometry(str(round(x * a)) + 'x' + str(round(y * b)) + '+' + str(round((-a * x + x) / 2)) + '+' + str(
            round((-b * y + y) / 2)))
        self.title('Invoice Refund')

        self.heading = tk.Label(self, text='Refund', font='Courier 25 bold')
        self.heading.grid(row=0, column=1)

        self.refund_type_variable = tk.StringVar(self)
        self.refund_type_variable.set(None)
        cash_button = tk.Radiobutton(self, text='Cash Refund',
                                     variable=self.refund_type_variable, value='cash', anchor='s')
        cash_button.grid(row=1, column=0)

        transfer_button = tk.Radiobutton(self, text='Transfer Refund',
                                         variable=self.refund_type_variable, value='transfer', anchor='s')
        transfer_button.grid(row=1, column=2)

        self.cancel_button = tk.Button(self, text='Cancel', command=self.destroy,
                                       activebackground=button_pressed_color)
        self.cancel_button.grid(row=2, column=0, sticky='E')
        self.cancel_button.bind('<Enter>', on_hover)
        self.cancel_button.bind('<Leave>', off_hover)


        self.submit_button = tk.Button(self, text='Submit', command=self.submit_refund,
                                       activebackground = button_pressed_color)
        self.submit_button.grid(row=2, column=2, sticky='W')
        self.submit_button.bind('<Enter>', on_hover)
        self.submit_button.bind('<Leave>', off_hover)

    @activity_log
    def submit_refund(self):
        if self.databaseConnection != None:
            selected_refund_type = self.refund_type_variable.get()
            if selected_refund_type != 'None':
                answer = messagebox.askyesno('Refund',
                                             f'Are you sure you want to refund ${self.receipt_data["invoice_total"]}\n '
                                             f'to {self.receipt_data["fullname"]} via {selected_refund_type}?',
                                             parent=self)
                if answer:
                    current_refund_no = self.databaseConnection.query(f'select auto_increment from '
                                                                      f'information_schema.tables '
                                                                      f'where table_schema = "{DATABASE_NAME}" '
                                                                      f'and table_name = "refunds";')
                    if selected_refund_type == 'cash':
                        self.cash_refund = self.receipt_data['invoice_total']
                        self.transfer_refund = 0
                    else:
                        self.cash_refund = 0
                        self.transfer_refund = self.receipt_data['invoice_total']

                    current_refund_no = current_refund_no[0][0]

                    today = datetime.today().strftime(format='%d-%m-%Y')
                    refund_pdf = RefundGenerator(pdf_name=f'{current_refund_no}.pdf',
                                                 invoice_total=self.receipt_data['invoice_total'],
                                                 cash_paid=self.receipt_data['cash_amount'],
                                                 transfer_paid=self.receipt_data['transfer_amount'],
                                                 amount_paid=self.receipt_data['invoice_total'],
                                                 amount_refunded=self.receipt_data['invoice_total'],
                                                 cash_or_transfer=selected_refund_type.capitalize())

                    item_codes = [c[0] for c in self.receipt_line]
                    item_desc = [c[1] for c in self.receipt_line]
                    item_qty = [c[3] for c in self.receipt_line]
                    item_value = [c[2] for c in self.receipt_line]
                    item_subtotal = [c[4] for c in self.receipt_line]

                    refund_pdf.invoice_line(item_codes=item_codes,
                                            item_descs=item_desc,
                                            item_prices=item_value,
                                            qtys=item_qty,
                                            subtotals=item_subtotal)
                    refund_pdf.memberName(self.receipt_data['fullname'])
                    refund_pdf.refundNo(current_refund_no)
                    refund_pdf.refundDate(today)
                    refund_pdf.streetAdress(self.receipt_data['street_address'])
                    refund_pdf.cityStatePostCode(self.receipt_data['suburb'],
                                                 self.receipt_data['state'],
                                                 self.receipt_data['postcode'])

                    error = 0
                    try:
                        self.databaseConnection.insert(f'insert into refunds ('
                                                       f'invoice_receipt_no, '
                                                       f'cash_amount, '
                                                       f'transfer_amount, '
                                                       f'payment_datetime) values ('
                                                       f'{self.receipt_data["receipt_no"]}, '
                                                       f'{self.cash_refund}, '
                                                       f'{self.transfer_refund}, '
                                                       f'now())')
                    except Exception as e:
                        self.databaseConnection.rollback()
                        print('Error Logged')
                        logger.exception(e)
                        error = 1
                        messagebox.showwarning('Unknown Error',
                                               f'An unknown error occurred and has been logged. Report to developer.',
                                               parent=self)

                    if error == 0:
                        self.databaseConnection.commit()
                        refund_pdf.save(f'{FOLDER}/refund_pdfs/')
                        self.destroy()
                        self.history_window.destroy()
                        self.main_menu.update_tables()
                        messagebox.showinfo('Refund Complete', 'Refund submitted successfully.', parent=self.main_menu)

            else:
                messagebox.showwarning('Refund Type', 'Please select a refund type.', parent=self)
        else:
            not_connected_message(self)


class Receipt_window(tk.Tk):
    def __init__(self, invoice_no, connection, main_menu, email_address, email_password, email_host, email_port, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)
        self.invoice_no = invoice_no
        self.report_callback_exception = log_unhandled_exception
        self.main_menu = main_menu
        self.databaseConnection = connection
        self.email_address = email_address
        self.email_password = email_password
        self.email_host = email_host
        self.email_port = email_port
        try:
            self.invoice_lines = self.databaseConnection.query(f'select '
                                                          f'item.item_code, '
                                                          f'item_description, '
                                                          f'invoice_item_value, '
                                                          f'item_qty, '
                                                          f'invoice_item_value*item_qty  '
                                                          f'from invoice_line '
                                                          f'join '
                                                          f'item '
                                                          f'on '
                                                          f'invoice_line.item_code = item.item_code '
                                                          f'where invoice_no = {invoice_no};')
        except:
            print('error 1')
            self.destroy()

        try:
            self.invoice_data = self.databaseConnection.query(f'select '
                                                              f'invoice_date, '
                                                              f'date_format(invoice_duedate, "%d-%m-%y"), '
                                                              f'invoice_total, '
                                                              f'concat_ws(" ", member_fname, member_lname), '
                                                              f'email, '
                                                              f'invoice_sent '
                                                              f'from invoice '
                                                              f'join members '
                                                              f'on '
                                                              f'invoice.member_no = members.member_no '
                                                              f'where '
                                                              f'invoice.invoice_no = {invoice_no};')
        except:
            print('error: 2')
            self.destroy()

        self.invoice_data = self.invoice_data[0]
        self.email = self.invoice_data[-2]
        self.invoice_duedate = self.invoice_data[1]
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure(2, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self.grid_rowconfigure(2, weight=1)
        self.grid_rowconfigure(3, weight=1)
        self.grid_rowconfigure(4, weight=1)
        x = self.winfo_screenwidth()
        y = self.winfo_screenheight()
        a = 0.45
        b = 0.5
        self.geometry(str(round(x * a)) + 'x' + str(round(y * b)) + '+' + str(round((-a * x + x) / 2)) + '+' + str(round((-b * y + y) / 2)))

        self.title('Unpaid Invoice')

        invoice_table(self, row=4, columnspan=3)

        for row in self.invoice_lines:
            self.invoice_table.insert('', 'end', text=row[0], values=row[1:])

        self.invoice_heading = tk.Label(self, text=f'Invoice: {self.invoice_no}', font='Courier 40 bold')
        self.invoice_heading.grid(row=0, columnspan=3, sticky='w')

        self.member_name_label = tk.Label(self, text=f'Member Name: {self.invoice_data[3]}', font='Courier 15 bold')
        self.member_name_label.grid(row=1, columnspan=3, sticky='nw', padx=10)

        self.issue_date_label = tk.Label(self, text=f'Issue Date: {self.invoice_data[0]} ', font='Courier 15 bold')
        self.issue_date_label.grid(row=2, columnspan=3, sticky='nw', padx=10)

        if self.invoice_data[-1] == 'No':
            self.send_invoice_button = tk.Button(self, text='Send Invoice', command=self.send_invoice,
                                                 activebackground = button_pressed_color)
            button_font = font.Font(family='Courier', size=15, weight='bold')
            self.send_invoice_button['font'] = button_font
            self.send_invoice_button.grid(row=3, column=2)
            self.send_invoice_button.bind('<Enter>', on_hover)
            self.send_invoice_button.bind('<Leave>', off_hover)
        else:
            self.send_invoice_button = tk.Button(self, text='Resend Invoice', command=self.send_invoice,
                                                 activebackground = button_pressed_color)
            button_font = font.Font(family='Courier', size=15, weight='bold')
            self.send_invoice_button['font'] = button_font
            self.send_invoice_button.grid(row=3, column=2)
            self.send_invoice_button.bind('<Enter>', on_hover)
            self.send_invoice_button.bind('<Leave>', off_hover)

        self.invoice_sent_label = tk.Label(self, text=f'Invoice Sent: {self.invoice_data[-1]} ', font='Courier 15 bold')
        self.invoice_sent_label.grid(row=3, columnspan=3, sticky='nw', padx=10)

        self.total_label = tk.Label(self, text=f'Total: ${self.invoice_data[2]} ', font='Courier 15 bold')
        self.total_label.grid(row=5, column=2, sticky='w')

        button_font = font.Font(family='Courier', size=20, weight='bold')

        self.delete_invoice_button = tk.Button(self, text='Delete Invoice', command=self.delete_invoice, width=18,
                                               activebackground = button_pressed_color)
        self.delete_invoice_button['font'] = button_font
        self.delete_invoice_button.grid(row=5, column=0)
        self.delete_invoice_button.bind('<Enter>', on_hover)
        self.delete_invoice_button.bind('<Leave>', off_hover)

        self.issue_receipt_button = tk.Button(self, text='Issue Receipt',
                                              command=lambda: ReceiptCashOrTransfer(self.invoice_data[2], self.invoice_no, connection, self, self.main_menu, self.email_address, self.email_password, self.email_host, self.email_port),
                                              width=18,
                                              activebackground = button_pressed_color)
        self.issue_receipt_button['font'] = button_font
        self.issue_receipt_button.grid(row=5, column=1, padx=10)
        self.issue_receipt_button.bind('<Enter>', on_hover)
        self.issue_receipt_button.bind('<Leave>', off_hover)

        self.attributes("-topmost", True)

    @activity_log
    def delete_invoice(self):
        # self.databaseConnection.reconnect()
        warning = messagebox.askyesno('Delete Invoice', 'Are you sure you want to delete this invoice?', parent=self)
        if warning:
            try:
                self.databaseConnection.insert(f'delete from invoice_line where invoice_no = {self.invoice_no}')
                self.databaseConnection.insert(f'delete from invoice where invoice_no = {self.invoice_no}')
                self.databaseConnection.commit()
            except Exception as e:
                self.databaseConnection.rollback()
                logger.exception(e)
                messagebox.showwarning('Unknown Error',
                                       f'An unknown error occurred and has been logged. Report to developer.',
                                       parent=self)
            else:
                self.main_menu.update_tables()
                self.destroy()

    @activity_log
    def send_invoice(self):
        email_yesno = messagebox.askyesno('Email Invoices', f'Would you like to send out this invoice to \n{self.invoice_data[-2]}?',
                                          parent=self)
        if email_yesno:
            print(f'sending invoice to {self.email}')
            sender = Emailer('Greater Western 4x4 Club Invoice',
                    f'To {self.invoice_data[3]},\n\n'
                    f' See attached your Greater Western 4x4 Invoice.\n\n'
                    f'Invoice Due Date: {self.invoice_duedate}\n'
                    f'Invoice Total: ${self.invoice_data[2]} ', self.email_address, self.email, self.email_password, self.email_host, self.email_port,
                             attachment_path=f'{FOLDER}/invoice_pdfs/{self.invoice_no}.pdf', filename=f'{self.invoice_no}.pdf')
            status = sender.get_status()
            print(status)
            if status == '1':
                self.databaseConnection.insert(f'update invoice set invoice_sent = "Yes" where invoice_no = {self.invoice_no}')
                self.databaseConnection.commit()
                self.main_menu.update_tables()
                messagebox.showinfo('Emailed Invoices', 'Invoice sent successfully', parent=self)
                self.destroy()
                Receipt_window(self.invoice_no, self.databaseConnection, self.main_menu, self.email_address, self.email_password, self.email_host, self.email_port)
            else:
                logger.exception(sender.return_error())
                messagebox.showerror('Email Failed', 'Invoice failed to be sent. Check email \ncredentials are correct.', parent=self)

        else:
            print('Cancelled')


class ReceiptCashOrTransfer(tk.Tk):
    @activity_log
    def __init__(self, invoice_total, invoice_no, connection, receipt_window, main_menu, email, password, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)
        self.grid_rowconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self.grid_rowconfigure(2, weight=1)
        self.grid_rowconfigure(3, weight=1)
        self.grid_rowconfigure(4, weight=1)
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        self.title('Issue Receipt')
        self.invoice_no = invoice_no
        self.invoice_total = invoice_total
        self.report_callback_exception = log_unhandled_exception
        self.receipt_window = receipt_window
        self.email_address = email
        self.email_password = password
        self.email_host = EMAIL_HOST
        self.email_port = EMAIL_PORT
        self.main_menu = main_menu
        self.databaseConnection = connection
        self.geometry(str(round(self.winfo_screenwidth() * 0.38)) + 'x' + str(round(self.winfo_screenheight() * 0.18)))
        self.invoice_total = invoice_total
        self.heading = tk.Label(self, text=f'Invoice Total: ${self.invoice_total}', font='Courier 25 bold')
        self.heading.grid(row=0, columnspan=2, padx=5)

        self.cash_label = tk.Label(self, text='Cash Amount:',  font='Courier 15')
        self.cash_label.grid(row=1, column=0)
        self.cash_var = tk.StringVar(self)
        self.cash_entry = tk.Entry(self, textvariable=self.cash_var, width=10)
        self.cash_entry.grid(row=1, column=1)

        self.transfer_label = tk.Label(self, text='Transfer Amount:',  font='Courier 15')
        self.transfer_label.grid(row=2, column=0)
        self.transfer_var = tk.StringVar(self)
        self.transfer_entry = tk.Entry(self, textvariable=self.transfer_var, width=10)
        self.transfer_entry.grid(row=2, column=1)

        self.date_label = tk.Label(self, text='Payment Date:', font='Courier 15')
        self.date_label.grid(row=3, column=0)
        self.date_calendar = DateEntry(self, date_pattern='dd/mm/yyyy')
        self.date_calendar.grid(row=3, column=1)

        self.attributes("-topmost", True)

        button_font = font.Font(family='Courier', size=20, weight='bold')

        self.cancel_button = tk.Button(self, text='Cancel', command=self.cancel_receipt,
                                       activebackground = button_pressed_color)
        self.cancel_button['font'] = button_font
        self.cancel_button.grid(row=4, column=0, pady=5)
        self.cancel_button.bind('<Enter>', on_hover)
        self.cancel_button.bind('<Leave>', off_hover)

        self.submit_button = tk.Button(self, text='Submit', command=self.submit_receipt,
                                       activebackground = button_pressed_color)
        self.submit_button['font'] = button_font
        self.submit_button.grid(row=4, column=1, pady=5)
        self.submit_button.bind('<Enter>', on_hover)
        self.submit_button.bind('<Leave>', off_hover)

    def cancel_receipt(self):
        self.destroy()

    @activity_log
    def submit_receipt(self):
        self.cash = self.cash_var.get()
        self.transfer = self.transfer_var.get()
        input_test_results = input_check(self.cash, self.transfer)
        if input_test_results != True:
            messagebox.showwarning('Input Error', input_test_results, parent=self)
        else:
            if self.cash == '':
                self.cash = 0
            if self.transfer == '':
                self.transfer = 0
            if self.transfer == 0 and self.cash == 0:
                messagebox.showwarning('Input Error', 'Cash and Transfer values cannot both be zero.')
            else:
                self.cash = float(self.cash)
                self.transfer = float(self.transfer)
                self.total = round(self.cash + self.transfer, 2)
                print(self.cash, self.transfer, self.total)
                if self.total != float(self.invoice_total):
                    messagebox.showerror('Value Error', 'Cash and transfer values do not equal invoice total.', parent=self)
                else:
                    self.receipt_insert()
                    self.destroy()
                    self.receipt_window.destroy()
                    messagebox.showinfo('Receipt Submitted', 'Receipt has successfully been submitted', parent=self.main_menu)
                    yesno = messagebox.askyesno('Send Receipt', f'Do you want to send this receipt to {self.member_email}? ')
                    if yesno:
                        self.send_email()

    @activity_log
    def receipt_insert(self):
        date = self.date_calendar.get()
        data = self.databaseConnection.query(f'select '
                                             f'invoice_total, '
                                             f'concat_ws(" ", member_fname, member_lname), '
                                             f'street_address, '
                                             f'suburb, state, '
                                             f'postcode, email '
                                             f'from '
                                             f'invoice '
                                             f'join '
                                             f'members '
                                             f'on '
                                             f'invoice.member_no = members.member_no '
                                             f'where '
                                             f'invoice_no = {self.invoice_no};')

        lines = self.databaseConnection.query(f'select '
                                              f'item.item_code, '
                                              f'item_description, '
                                              f'item_qty, '
                                              f'invoice_item_value, '
                                              f'item_qty*invoice_item_value '
                                              f'from invoice_line '
                                              f'join '
                                              f'item on '
                                              f'invoice_line.item_code = item.item_code '
                                              f'where invoice_no = {self.invoice_no};')
        invoice_total = data[0][0]
        self.fullname = data[0][1]
        street_address = data[0][2]
        suburb = data[0][3]
        state = data[0][4]
        postcode = data[0][5]
        self.member_email = data[0][6]
        receipt_date = datetime.today().date().strftime("%d-%m-%y")
        item_codes = [c[0] for c in lines]
        item_desc = [c[1] for c in lines]
        item_qty = [c[2] for c in lines]
        item_value = [c[3] for c in lines]
        item_subtotal = [c[4] for c in lines]

        self.current_receipt_no = self.databaseConnection.query('select max(invoice_receipt_no)+1 from invoice_receipt;')
        self.current_receipt_no = self.current_receipt_no[0][0]
        receipt_gen = ReceiptGenerator(f'R{self.current_receipt_no}.pdf', invoice_total, self.cash, self.transfer, invoice_total)
        # item_codes, desc, prices, qtys, subtotals
        receipt_gen.invoice_line(item_codes, item_desc, item_value, item_qty, item_subtotal)
        receipt_gen.memberName(self.fullname)
        receipt_gen.receiptNo(self.current_receipt_no)
        receipt_gen.receiptDate(receipt_date)
        receipt_gen.streetAdress(street_address)
        receipt_gen.cityStatePostCode(suburb, state, postcode)


        if date == datetime.today().strftime("%d/%m/%Y"):
            self.databaseConnection.insert(f'insert into invoice_receipt '
                                           f'(invoice_receipt_no, invoice_no, cash_amount, transfer_amount, payment_datetime, receipt_sent) '
                                           f'values '
                                           f'({self.current_receipt_no}, {self.invoice_no}, {self.cash}, {self.transfer}, now(), "No")')
        else:
            self.databaseConnection.insert(f'insert into invoice_receipt '
                                           f'(invoice_receipt_no, invoice_no, cash_amount, transfer_amount, payment_datetime, receipt_sent) '
                                           f'values '
                                           f'({self.current_receipt_no}, {self.invoice_no}, {self.cash}, {self.transfer}, str_to_date("{date}","%d/%m/%Y"), "No")')
        self.databaseConnection.commit()
        self.main_menu.update_tables()
        receipt_gen.save(f'{FOLDER}\\receipt_pdfs')
        print('Success')

    @activity_log
    def send_email(self):
        print(f'sending receipt to {self.member_email}')
        sender = Emailer('Greater Western 4x4 Club Receipt',  # subject
                f'To {self.fullname},\n\n'  # body
                f' See attached your Greater Western 4x4 receipt.\n\n',
                self.email_address,  # sender
                self.member_email,  # receiver
                self.email_password, self.email_host, self.email_port,  # email password
                attachment_path=f'{FOLDER}/receipt_pdfs/R{self.current_receipt_no}.pdf', filename=f'R{self.current_receipt_no}.pdf')  # receipt attachment
        status = sender.get_status()
        if status == '1':
            self.databaseConnection.insert(
                f'update invoice_receipt set receipt_sent = "Yes" where invoice_receipt_no = {self.current_receipt_no}')
            messagebox.showinfo('Receipt Sent', 'Receipt successfully sent.')
        else:
            logger.exception(sender.return_error())
            messagebox.showerror('Email Failed',
                                 'Receipt failed to be sent. Check email \ncredentials are correct.',
                                 parent=self.main_menu)


class ExpenseWindow(tk.Tk):
    @activity_log
    def __init__(self, connection, main_menu, *args, **kwargs):
        self.main_menu = main_menu
        tk.Tk.__init__(self, *args, **kwargs)
        self.title('Add Expense')
        self.databaseConnection = connection
        self.report_callback_exception = log_unhandled_exception
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure(2, weight=1)
        self.grid_columnconfigure(3, weight=1)
        self.grid_columnconfigure(4, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self.grid_rowconfigure(2, weight=1)
        self.grid_rowconfigure(3, weight=1)
        self.grid_rowconfigure(4, weight=1)
        self.grid_rowconfigure(5, weight=1)
        self.grid_rowconfigure(6, weight=1)
        self.grid_rowconfigure(7, weight=1)

        x = self.winfo_screenwidth()
        y = self.winfo_screenheight()
        a = 0.3
        b = 0.35
        self.geometry(str(round(x * a)) +
                      'x'
                      + str(round(y * b)) +
                      '+'
                      + str(round((-a * x + x) / 2)) +
                      '+'
                      + str(round((-b * y + y) / 2)))

        if self.databaseConnection != None:
            self.expenses = self.databaseConnection.query(
                "select expense_description, expense_id from expense where expense_id != 1 and hidden = 0;")
        else:
            self.expenses = []

        self.expense_var = tk.StringVar(self)
        self.expense_var.set('Select Category')

        self.expenses_desc = [c[0] for c in self.expenses]
        self.expenses_desc.append('---- New Category ----')

        self.expense_dict = dict(self.expenses)

        self.expense_dropdown = tk.OptionMenu(self, self.expense_var, *self.expenses_desc)
        self.expense_dropdown.grid(row=0, columnspan=2)

        self.cash_label = tk.Label(self, text='Cash Amount:')
        self.cash_label.grid(row=2, column=0)

        self.expense_cash_var = tk.StringVar(self)
        self.expense_cash = tk.Entry(self, textvariable=self.expense_cash_var, width=10)
        self.expense_cash.grid(row=2, column=1)

        self.cash_label = tk.Label(self, text='Transfer Amount:')
        self.cash_label.grid(row=3, column=0)

        self.expense_transfer_var = tk.StringVar(self)
        self.expense_transfer = tk.Entry(self, textvariable=self.expense_transfer_var, width=10)
        self.expense_transfer.grid(row=3, column=1)

        self.date_label = tk.Label(self, text='Payment Date:')
        self.date_label.grid(row=4, column=0)
        self.date_calendar = DateEntry(self, date_pattern='dd/mm/yyyy')
        self.date_calendar.grid(row=4, column=1)

        self.note_label = tk.Label(self, text='Notes')
        self.note_label.grid(row=5, columnspan=2)

        self.note = tk.Text(self, width=50, height=7)
        self.note.grid(row=6, columnspan=2, padx=10, pady=10)

        cancel = tk.Button(self, text='Cancel', command=self.destroy,
                           activebackground = button_pressed_color)
        cancel.grid(row=7, column=0, pady=10)
        cancel.bind('<Enter>', on_hover)
        cancel.bind('<Leave>', off_hover)

        button = tk.Button(self, text='Submit', command=self.submit,
                           activebackground = button_pressed_color)
        button.grid(row=7, column=1, pady=10)
        button.bind('<Enter>', on_hover)
        button.bind('<Leave>', off_hover)

        self.expense_var.trace('w', self.create_new_category)
        self.new_cat_exists = 0
        self.new_category_var = tk.StringVar(self)

    @activity_log
    def create_new_category(self, *args):
        expense = self.expense_var.get()
        if expense == '---- New Category ----':
            if self.new_cat_exists == 0:  # doesnt exist
                self.new_category_label = tk.Label(self, text='New Category:')
                self.new_category_label.grid(row=1, column=0)
                self.new_category_entry = tk.Entry(self, textvariable=self.new_category_var)
                self.new_category_entry.grid(row=1, column=1)
                self.new_cat_exists = 1  # exists
        else:
            if self.new_cat_exists == 1:
                self.new_category_label.grid_forget()
                self.new_category_entry.grid_forget()
                self.new_category_var.set('')
                self.new_cat_exists = 0  # set to doesnt exist

    @activity_log
    def submit(self):
        if self.databaseConnection != None:
            date = self.date_calendar.get()
            current_expense_no = self.databaseConnection.query('select max(expense_receipt_no)+1 from expense_receipt;')
            current_expense_no = current_expense_no[0][0]
            notes = self.get_text()
            self.cash = self.expense_cash_var.get()
            self.transfer = self.expense_transfer_var.get()

            if self.expense_var.get() == 'Select Category':
                messagebox.showerror('Select Category', 'Please select a category in the drop down menu\n'
                                                        'or add a new one.', parent=self)
            elif self.expense_var.get() == '---- New Category ----' and len(self.new_category_var.get()) == 0:
                    messagebox.showerror('New Category', 'New Category field cannot be empty', parent=self)
            elif (self.cash == '' or self.cash == 0) and (self.transfer == '' or self.transfer == 0) :
                messagebox.showerror('Missing Values', 'Cash and transfer values cannot be both zero', parent=self)
            else:
                input_test_results = input_check(self.cash, self.transfer)
                if input_test_results != True:
                    messagebox.showwarning('Input Error', input_test_results, parent=self)
                else:
                    if self.cash == '':
                        self.cash = 0
                    if self.transfer == '':
                        self.transfer = 0
                    self.cash = float(self.cash)
                    self.transfer = float(self.transfer)
                    try:
                        if self.expense_var.get() == '---- New Category ----':
                            self.insert_new_cat()
                            self.expense_id = self.current_expense_cat_no
                        else:
                            self.expense_id = self.expense_dict[self.expense_var.get()]
                        if date == datetime.today().strftime(format="%d/%m/%Y"):
                            date_field = 'now()'
                        else:
                            date_field = f'str_to_date("{date}","%d/%m/%Y")'
                        self.databaseConnection.insert(f'insert into '
                                                       f'expense_receipt '
                                                       f'(expense_receipt_no, '
                                                       f'expense_id, '
                                                       f'cash_amount, '
                                                       f'transfer_amount, '
                                                       f'payment_datetime, '
                                                       f'expense_notes) '
                                                       f'values '
                                                       f'({current_expense_no}, '
                                                       f'{self.expense_id}, '
                                                       f'-{self.cash}, '  # make negative as is an expense
                                                       f'-{self.transfer}, '   # make negative as is an expense
                                                       f'{date_field}, '
                                                       f'"{notes}");')
                        self.databaseConnection.commit()
                    except Exception as e:
                        self.databaseConnection.rollback()
                        logger.exception(e)
                        messagebox.showwarning('Unknown Error',
                                               f'An unknown error occurred and has been logged. Report to developer.',
                                               parent=self)
                    else:
                        self.main_menu.update_tables()
                        self.destroy()
                        messagebox.showinfo('Expense Recorded', 'New expense recorded successfully',
                                            parent=self.main_menu)

    def get_text(self):
        text = self.note.get('1.0', 'end')
        return text

    def insert_new_cat(self):
        self.current_expense_cat_no = self.databaseConnection.query('select '
                                                       'max(expense_id)+1 '
                                                       'from expense;')
        self.current_expense_cat_no = self.current_expense_cat_no[0][0]
        self.databaseConnection.insert(f'insert into '
                                       f'expense '
                                       f'(expense_id, '
                                       f'expense_description) '
                                       f'values '
                                       f'({self.current_expense_cat_no}, '
                                       f'"{self.databaseConnection.sanitise(self.new_category_var.get())}");')
        # self.databaseConnection.commit()


class IncomeWindow(tk.Tk):
    @activity_log
    def __init__(self, connection, main_menu, *args, **kwargs):
        self.main_menu = main_menu
        self.report_callback_exception = log_unhandled_exception
        tk.Tk.__init__(self, *args, **kwargs)
        self.databaseConnection = connection
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure(2, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self.grid_rowconfigure(2, weight=1)
        self.grid_rowconfigure(3, weight=1)
        self.grid_rowconfigure(4, weight=1)
        self.grid_rowconfigure(5, weight=1)
        self.grid_rowconfigure(6, weight=1)
        self.grid_rowconfigure(7, weight=1)

        x = self.winfo_screenwidth()
        y = self.winfo_screenheight()
        a = 0.3
        b = 0.35
        self.geometry(str(round(x * a)) +
                      'x'
                      + str(round(y * b)) +
                      '+'
                      + str(round((-a * x + x) / 2)) +
                      '+'
                      + str(round((-b * y + y) / 2)))

        self.title('Add Income')
        if self.databaseConnection != None:
            self.incomes = self.databaseConnection.query(
                "select income_description, income_id from income where income_id != 1 and hidden = 0;")
        else:
            self.incomes = []

        self.income_var = tk.StringVar(self)
        self.income_var.set('Select Category')

        self.incomes_desc = [c[0] for c in self.incomes]
        self.incomes_desc.append('---- New Category ----')

        self.income_dict = dict(self.incomes)

        self.income_dropdown = tk.OptionMenu(self, self.income_var, *self.incomes_desc)
        self.income_dropdown.grid(row=0, columnspan=2)

        self.cash_label = tk.Label(self, text='Cash Amount:')
        self.cash_label.grid(row=2, column=0)

        self.income_cash_var = tk.StringVar(self)
        self.income_cash = tk.Entry(self, textvariable=self.income_cash_var, width=10)
        self.income_cash.grid(row=2, column=1)

        self.cash_label = tk.Label(self, text='Transfer Amount:')
        self.cash_label.grid(row=3, column=0)

        self.income_transfer_var = tk.StringVar(self)
        self.income_transfer = tk.Entry(self, textvariable=self.income_transfer_var, width=10)
        self.income_transfer.grid(row=3, column=1)

        self.date_label = tk.Label(self, text='Payment Date:')
        self.date_label.grid(row=4, column=0)
        self.date_calendar = DateEntry(self, date_pattern='dd/mm/yyyy')
        self.date_calendar.grid(row=4, column=1)

        self.note_label = tk.Label(self, text='Notes')
        self.note_label.grid(row=5, columnspan=2)

        self.note = tk.Text(self, width=50, height=7)
        self.note.grid(row=6, columnspan=2, padx=10, pady=10)

        cancel = tk.Button(self, text='Cancel', command=self.destroy,
                           activebackground=button_pressed_color
                           )
        cancel.grid(row=7, column=0, pady=10)
        cancel.bind('<Enter>', on_hover)
        cancel.bind('<Leave>', off_hover)

        button = tk.Button(self, text='Submit', command=self.submit,
                           activebackground=button_pressed_color
                           )
        button.grid(row=7, column=1, pady=10)
        button.bind('<Enter>', on_hover)
        button.bind('<Leave>', off_hover)

        self.income_var.trace('w', self.create_new_category)
        self.new_cat_exists = 0
        self.new_category_var = tk.StringVar(self)

    @activity_log
    def create_new_category(self, *args):
        income = self.income_var.get()
        if income == '---- New Category ----':
            if self.new_cat_exists == 0:  # doesnt exist
                self.new_category_label = tk.Label(self, text='New Category:')
                self.new_category_label.grid(row=1, column=0)
                self.new_category_entry = tk.Entry(self, textvariable=self.new_category_var)
                self.new_category_entry.grid(row=1, column=1)
                self.new_cat_exists = 1  # exists

        else:
            if self.new_cat_exists == 1:
                self.new_category_label.grid_forget()
                self.new_category_entry.grid_forget()
                self.new_category_var.set('')
                self.new_cat_exists = 0  # set to doesnt exist

    @activity_log
    def submit(self):
        if self.databaseConnection != None:
            date = self.date_calendar.get()
            print(date)
            current_income_no = self.databaseConnection.query('select max(income_receipt_no)+1 from income_receipt;')
            current_income_no = current_income_no[0][0]
            notes = self.get_text()
            self.cash = self.income_cash_var.get()
            self.transfer = self.income_transfer_var.get()

            if self.income_var.get() == 'Select Category':
                messagebox.showerror('Select Category', 'Please select a category in the drop down menu\n'
                                                        'or add a new one.', parent=self)
            elif self.income_var.get() == '---- New Category ----' and len(self.new_category_var.get()) == 0:
                    messagebox.showerror('New Category', 'New Category field cannot be empty', parent=self)
            elif (self.cash == '' or self.cash == 0) and (self.transfer == '' or self.transfer == 0) :
                messagebox.showerror('Missing Values', 'Cash and transfer values cannot be both zero', parent=self)
            else:
                input_test_results = input_check(self.cash, self.transfer)
                if input_test_results != True:
                    messagebox.showwarning('Input Error', input_test_results, parent=self)
                else:
                    if self.cash == '':
                        self.cash = 0
                    if self.transfer == '':
                        self.transfer = 0
                    self.cash = float(self.cash)
                    self.transfer = float(self.transfer)
                    if self.income_var.get() == '---- New Category ----':
                        self.insert_new_cat()
                        self.income_id = self.current_income_cat_no
                    else:
                        self.income_id = self.income_dict[self.income_var.get()]
                    if date == datetime.today().strftime(format="%d/%m/%Y"):
                        date_field = 'now()'
                    else:
                        date_field = f'str_to_date("{date}","%d/%m/%Y")'
                    try:
                        self.databaseConnection.insert(f'insert into '
                                                       f'income_receipt '
                                                       f'(income_receipt_no, '
                                                       f'income_id, '
                                                       f'cash_amount, '
                                                       f'transfer_amount, '
                                                       f'payment_datetime, '
                                                       f'income_notes) '
                                                       f'values '
                                                       f'({current_income_no}, '
                                                       f'{self.income_id}, '
                                                       f'{self.cash}, '  
                                                       f'{self.transfer}, '   
                                                       f'{date_field}, '
                                                       f'"{notes}");')
                    except Exception as e:
                        self.databaseConnection.rollback()
                        logger.exception(e)
                        messagebox.showwarning('Unknown Error',
                                               f'An unknown error occurred and has been logged. Report to developer.',
                                               parent=self)
                    else:
                        self.databaseConnection.commit()
                        self.main_menu.update_tables()
                        self.destroy()
                        messagebox.showinfo('Income Recorded', 'New income recorded successfully',
                                            parent=self.main_menu)


    def get_text(self):
        text = self.note.get('1.0', 'end')
        return text

    def insert_new_cat(self):
        self.current_income_cat_no = self.databaseConnection.query('select '
                                                       'max(income_id)+1 '
                                                       'from income;')
        self.current_income_cat_no = self.current_income_cat_no[0][0]
        self.databaseConnection.insert(f'insert into '
                                       f'income '
                                       f'(income_id, '
                                       f'income_description) '
                                       f'values '
                                       f'({self.current_income_cat_no}, '
                                       f'"{self.databaseConnection.sanitise(self.new_category_var.get())}");')
        # self.databaseConnection.commit()


class AutoInvoicing(tk.Tk):
    @activity_log
    def __init__(self, connection, main_menu, email_address, email_password, email_host, email_port, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)
        self.main_menu = main_menu
        self.databaseConnection = connection
        self.report_callback_exception = log_unhandled_exception
        self.email_address = email_address
        self.email_password = email_password
        self.email_host = email_host
        self.email_port = email_port
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=2)
        self.grid_rowconfigure(2, weight=2)
        x = self.winfo_screenwidth()
        y = self.winfo_screenheight()
        a = 0.20
        b = 0.20
        self.title('Auto Invoicer')

        self.geometry(str(round(x * a)) +
                      'x'
                      + str(round(y * b)) +
                      '+'
                      + str(round((-a*x + x)/2)) +
                      '+'
                      + str(round((-b*y + y)/2)))
        if self.databaseConnection != None:
            self.active_members = self.databaseConnection.query('select '
                                                            'member_no, '
                                                            'concat_ws(" ", member_fname, member_lname), '
                                                            'street_address, '
                                                            'suburb, '
                                                            'state, '
                                                            'postcode, '
                                                            'email '
                                                            'from '
                                                            'members '
                                                            'where '
                                                            'member_status = "ACTIVE";')
            self.current_invoice_no = self.databaseConnection.query('select max(invoice_no)+1 from invoice')
            self.current_invoice_no = self.current_invoice_no[0][0]
            self.bank_details = self.databaseConnection.query(f'select bank_name, bsb, account_no from bank_detail')
            self.bank_details = self.bank_details[0]
        else:
            self.active_members = []
            self.current_invoice_no = []
            self.bank_details = []

        self.auto_label = tk.Label(self, text='Auto Invoicer', font='courier 15 bold')
        self.auto_label.grid(row=0, columnspan=2)
        self.due_date_label = tk.Label(self, text='Due Date:')
        self.due_date_label.grid(row=1, column=0, sticky='N')
        self.due_date_var = tk.StringVar(self)
        current_year = datetime.today().year
        self.calendar = DateEntry(self, date_pattern='dd/mm/yy')
        self.calendar.grid(row=1, column=1, sticky='N')
        if datetime.today() < datetime(current_year, 7, 1):
            current_year = str(datetime.today().year)
            self.calendar.set_date('01/07/'+current_year[-2:])
        else:
            current_year = datetime(current_year + 1, 7, 1).year
            current_year = str(current_year)
            self.calendar.set_date('01/07/' + current_year[-2:])
        self.membership_price_label = tk.Label(self, text='Membership Price :')
        self.membership_price_label.grid(row=2, column=0, sticky='N')
        self.membership_var = tk.StringVar(self)
        self.membership_var.set(100)
        self.membership_entry = tk.Entry(self, textvariable=self.membership_var, width=10)
        self.membership_entry.grid(row=2, column=1, sticky='N')

        self.cancel = tk.Button(self, text='Cancel', command=self.destroy,
                                activebackground = button_pressed_color)
        self.cancel.grid(row=3, column=0, sticky='N', pady=(0,10))
        self.cancel.bind('<Enter>', on_hover)
        self.cancel.bind('<Leave>', off_hover)

        self.start_button = tk.Button(self, text='Start', command=self.start,
                                      activebackground = button_pressed_color)
        self.start_button.grid(row=3, column=1, sticky='N', pady=(0,10))
        self.start_button.bind('<Enter>', on_hover)
        self.start_button.bind('<Leave>', off_hover)


    @activity_log
    def start(self):
        if self.databaseConnection != None:
            self.due_date = self.calendar.get()
            self.price = self.membership_var.get()
            if match(r'^-?\d+\.?(\d+)?$', self.price) is None:
                messagebox.showwarning('Input Error', 'Price must be numeric', parent=self)
            else:
                if float(self.price) <= 0:
                    messagebox.showwarning('Input Error', 'Price must be greater than zero', parent=self)
                else:
                    yesno = messagebox.askyesno('Auto Invoicer',
                                                'This program will produce invoices for all active members.'
                                                ' Are you sure you want to continue?', parent=self)
                    self.invoice_email_list = []
                    if yesno:
                        for member in self.active_members:
                            status = self.create_invoice(member)  # return False if aborted
                            if not status:
                                break
                        if status:
                            self.main_menu.update_tables()
                            email_yesno = messagebox.askyesno('Email Invoices', 'Would you like to send out these invoices via email?',
                                                          parent=self)
                            if email_yesno:
                                error = EmailProgress(self.invoice_email_list, self.databaseConnection, self.main_menu, self.email_address, self.email_password, self.email_host, self.email_port)
                                if error.get_error_status() == '1':
                                    messagebox.showinfo('Email Invoices', 'All invoices sent out', parent=self.main_menu)
                    self.destroy()
        else:
            not_connected_message(self)

    def create_invoice(self, member):
        try:
            invoice_filename = f'{self.current_invoice_no}.pdf'
            self.databaseConnection.insert(f'insert into invoice '
                                           f'(invoice_no, invoice_date, invoice_duedate,invoice_total, member_no, invoice_sent) values '
                                           f'({self.current_invoice_no}, now(), str_to_date("{self.due_date}","%d/%m/%y"),'
                                           f'{self.price},{member[0]}, "No")')
            self.databaseConnection.insert(f'insert into invoice_line '
                                           f'(invoice_no, item_code, item_qty, invoice_item_value) values '
                                           f'({self.current_invoice_no}, 1, 1,{self.price})')
            self.invoice_email_list.append((invoice_filename, member[-1], member[1]))
            self.current_invoice_no += 1
            invoice = InvoiceGenerator(invoice_filename, self.price, self.due_date)
            invoice.invoiceNo(f'{self.current_invoice_no}')
            invoice.BankDetails(self.bank_details[0], self.bank_details[1], self.bank_details[2])
            invoice.memberName(member[1])
            invoice.streetAdress(member[2])
            invoice.cityStatePostCode(member[3], member[4], member[5])
            invoice.invoiceDate(datetime.today().date().strftime('%d/%m/%y'))
            invoice.invoice_line([1], ['Annual Membership Renewal Fee'], [self.price], [1], [self.price])
            invoice.save(f'{FOLDER}\\invoice_pdfs')
        except self.databaseConnection.errors.InterfaceError:
            messagebox.showerror('Connection Lost', 'A connection with the database has been lost', parent=self)
            return False
        except Exception as e:
            self.databaseConnection.rollback()
            logging.exception(e)
            response = abortretryignore('Failed creating invoice',
                                        f'There was an error creating invoice {self.current_invoice_no} for '
                                        f'{member[0]}', parent=self)
            if response is None:  # ignore
                return True
            elif response:  # retry
                self.create_invoice(member)
            else:  # abort
                return False
        else:
            self.databaseConnection.commit()
            return True


class EmailProgress(tk.Tk):
    @activity_log
    def __init__(self, email_list, connection, main_menu, email_address, email_password, email_host, email_port, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)
        self.databaseConnection = connection
        self.report_callback_exception = log_unhandled_exception
        self.main_menu = main_menu
        self.email_address = email_address
        self.email_password = email_password
        self.email_host = email_host
        self.title('Sending Invoices')
        self.email_port = email_port
        x = self.winfo_screenwidth()
        y = self.winfo_screenheight()
        self.geometry(str(round(x * 0.3)) + 'x' + str(round(y * 0.2)) + '+' + str(round((-0.3*x + x)/2)) + '+' + str(round((-0.2*y + y)/2)))
        self.progress = Progressbar(self, orient=tk.HORIZONTAL, length=300, mode='determinate')
        progress_label = tk.Label(self, text='Sending Invoices...', font='courier 20 bold')
        progress_label.pack()
        self.progress.pack()
        current_email = tk.Label(self, text=f'Sending {email_list[0][0]} to {email_list[0][1]}', font='courier 10')
        current_email.pack()
        i = 1
        number_of_members = len(email_list)

        '''
        email_list contains:
        - invoice_filemame,
        - member_email,
        - member_fullname
        '''

        for email in email_list:
            sender = Emailer(subject='Greater Western 4x4 Club Membership Invoice',
                            body=f'To {email[-1]},\n\n'
                            'See attached your membership invoice',
                            sender=self.email_address,  # treasurer email (sender email)
                            receiver=email[1],  # recipients email
                            password=self.email_password,
                            email_host=self.email_host,
                            email_port=self.email_port,
                            attachment_path=f'{FOLDER}/invoice_pdfs/{email[0]}',
                            filename=f'{email[0]}')
            status = sender.get_status()
            if status == '1':
                invoice_no = email[0][:-4]
                self.databaseConnection.insert(f'update invoice set invoice_sent = "Yes" where invoice_no = {invoice_no}')
                self.databaseConnection.commit()

                current_value = (100/number_of_members)*i
                self.progresser(current_value)
                self.progress.update()
                i += 1
                # new_email_str = f'Sending {email[0]} to {email[1]}'
                new_email_str = f'Sending {email[0]} to testName@testmail.com'
                current_email.configure(text=new_email_str)
                self.update_idletasks()
                self.error = '1'
            else:
                logger.exception(sender.return_error())
                messagebox.showerror('Email Failed', 'Invoice failed to be sent. '
                                                     'Check email \ncredentials are correct.',
                                     parent=self.main_menu)
                self.error = '0'
                break
        self.main_menu.update_tables()
        self.destroy()

    @activity_log
    def progresser(self, current_value):
        self.progress['value'] = current_value

    def get_error_status(self):
        return self.error


class CashTransfers(tk.Tk):
    @activity_log
    def __init__(self, connection, main_menu, *args, **kwargs):
        self.main_menu = main_menu
        self.databaseConnection = connection
        self.report_callback_exception = log_unhandled_exception
        tk.Tk.__init__(self, *args, **kwargs)

        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure(2, weight=1)
        self.grid_columnconfigure(3, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self.grid_rowconfigure(2, weight=1)
        self.title('Transfer')
        x = self.winfo_screenwidth()
        y = self.winfo_screenheight()
        a = 0.4
        b = 0.2
        self.geometry(str(round(x * a)) +
                      'x'
                      + str(round(y * b)) +
                      '+'
                      + str(round((-a * x + x) / 2)) +
                      '+'
                      + str(round((-b * y + y) / 2)))
        if self.databaseConnection != None:
            self.balances = self.databaseConnection.query('select @cashSum, @bankSum')
            cash = self.balances[0][0]
            bank = self.balances[0][1]
        else:
            cash = 0
            bank = 0
        self.cash_label = tk.Label(self, text='Cash Balance', font='courier 15 bold')
        self.cash_label.grid(row=0, column=0)

        self.cash_balance_label = tk.Label(self, text=f'${cash}', font='courier 20 bold')
        self.cash_balance_label.grid(row=1, column=0)

        self.amount_label = tk.Label(self, text='Amount', font='courier 15 bold')
        self.amount_label.grid(row=0, column=1, columnspan=2)

        self.transfer_label = tk.Label(self, text='Bank Balance', font='courier 15 bold')
        self.transfer_label.grid(row=0, column=3)

        self.bank_balance_label = tk.Label(self, text=f'${bank}', font='courier 20 bold')
        self.bank_balance_label.grid(row=1, column=3)

        self.transfer_entry_var = tk.StringVar(self)
        self.transfer_entry = tk.Entry(self, textvariable=self.transfer_entry_var)
        self.transfer_entry.grid(row=1, column=1, columnspan=2)

        self.left_button = tk.Button(self, text='Cash Out', command=self.left_button,
                                     activebackground = button_pressed_color)
        self.left_button.grid(row=2, column=1)
        self.left_button.bind('<Enter>', on_hover)
        self.left_button.bind('<Leave>', off_hover)


        self.right_button = tk.Button(self, text='Deposit', command=self.right_button,
                                      activebackground = button_pressed_color)
        self.right_button.grid(row=2, column=2)
        self.right_button.bind('<Enter>', on_hover)
        self.right_button.bind('<Leave>', off_hover)

    @activity_log
    def left_button(self):
        if self.databaseConnection != None:
            transfer = self.transfer_entry_var.get()
            if transfer == 0 or transfer == '':
                messagebox.showwarning('Error', 'Transfer value cannot be zero', parent=self)
            else:
                if match(r'^-?\d+\.?(\d+)?$', transfer) is None:
                    messagebox.showwarning('Error', 'Transfer value must be numeric', parent=self)
                else:
                    if float(transfer) < 0:
                        messagebox.showwarning('Error', 'Transfer value cannot be less than zero', parent=self)
                    else:
                        current_transfer_no = self.databaseConnection.query('select max(transfer_no)+1 from transfer;')
                        self.databaseConnection.insert('insert into transfer '
                                                       '(transfer_no, cash_amount, transfer_amount, payment_datetime)'
                                                       'values '
                                                       f'({current_transfer_no[0][0]}, {transfer}, -{transfer}, now())')
                        self.databaseConnection.commit()
                        self.main_menu.update_tables()
                        self.balances = self.databaseConnection.query('select @cashSum, @bankSum')
                        cash = self.balances[0][0]
                        bank = self.balances[0][1]
                        self.cash_balance_label.configure(text=f'${cash}')
                        self.bank_balance_label.configure(text=f'${bank}')
                        self.transfer_entry_var.set('')
        else:
            not_connected_message(self)

    @activity_log
    def right_button(self):
        if self.databaseConnection != None:
            transfer = self.transfer_entry_var.get()
            if transfer == 0 or transfer == '':
                messagebox.showwarning('Error', 'Transfer value cannot be zero', parent=self)
            else:
                if match(r'^-?\d+\.?(\d+)?$', transfer) is None:
                    messagebox.showwarning('Error', 'Transfer value must be numeric', parent=self)
                else:
                    if float(transfer) < 0:
                        messagebox.showwarning('Error', 'Transfer value cannot be less than zero', parent=self)
                    else:
                        current_transfer_no = self.databaseConnection.query('select max(transfer_no)+1 from transfer;')
                        self.databaseConnection.insert('insert into transfer '
                                                       '(transfer_no, cash_amount, transfer_amount, payment_datetime)'
                                                       'values '
                                                       f'({current_transfer_no[0][0]}, -{transfer}, {transfer}, now())')
                        self.databaseConnection.commit()
                        self.main_menu.update_tables()
                        self.balances = self.databaseConnection.query('select @cashSum, @bankSum')
                        cash = self.balances[0][0]
                        bank = self.balances[0][1]
                        self.cash_balance_label.configure(text=f'${cash}')
                        self.bank_balance_label.configure(text=f'${bank}')
                        self.transfer_entry_var.set('')
        else:
            not_connected_message(self)


class ReportPeriod(tk.Tk):
    @activity_log
    def __init__(self, connection, main_menu, *args, **kwargs):
        self.databaseConnection = connection
        self.main_menu = main_menu
        self.report_callback_exception = log_unhandled_exception
        tk.Tk.__init__(self, *args, **kwargs)
        self.title('Committee Report')
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self.grid_rowconfigure(2, weight=1)
        self.grid_rowconfigure(3, weight=1)

        x = self.winfo_screenwidth()
        y = self.winfo_screenheight()
        a = 0.35  # width
        b = 0.2  # height
        self.geometry(str(round(x * a)) +
                      'x'
                      + str(round(y * b)) +
                      '+'
                      + str(round((-a * x + x) / 2)) +
                      '+'
                      + str(round((-b * y + y) / 2)))

        self.report_title = tk.Label(self, text='Committee Report', font='courier 30 bold')
        self.report_title.grid(row=0, columnspan=2)

        self.report_start_label = tk.Label(self, text='Report Start Date', font='courier 12')
        self.report_start_label.grid(row=1, column=0)

        self.start_calendar = DateEntry(self, date_pattern='dd/mm/yyyy')
        self.start_calendar.grid(row=1, column=1)

        self.report_end_label = tk.Label(self, text='Report End Date', font='courier 12')
        self.report_end_label.grid(row=2, column=0)

        self.end_calendar = DateEntry(self, date_pattern='dd/mm/yyyy')
        self.end_calendar.grid(row=2, column=1)

        self.cancel = tk.Button(self, text='Cancel', command=self.destroy,
                                activebackground = button_pressed_color)
        self.cancel.grid(row=3, column=0)
        self.cancel.bind('<Enter>', on_hover)
        self.cancel.bind('<Leave>', off_hover)

        self.create = tk.Button(self, text='Create Report', command=self.committee_report_print,
                                activebackground = button_pressed_color)
        self.create.grid(row=3, column=1)
        self.create.bind('<Enter>', on_hover)
        self.create.bind('<Leave>', off_hover)


    @activity_log
    def committee_report_print(self):
        if self.databaseConnection != None:
            report_start = self.start_calendar.get()
            report_end = self.end_calendar.get()

            report_start_flip = report_start[6:]+report_start[2:6]+report_start[:2]
            report_end_flip = report_end[6:] + report_end[2:6] + report_end[:2]

            # self.balances = self.databaseConnection.query('select @cashSum, @bankSum')
            self.cash_balance = self.databaseConnection.query(f"""
                select sum(cash_amount) from payment_history where payment_datetime <= '{report_end_flip}'
                """)[0][0]

            print(f'committee_cash_balance: {self.cash_balance}')


            self.bank_balance = self.databaseConnection.query(f"""
                select sum(transfer_amount) from payment_history where payment_datetime <= '{report_end_flip}'
            """)[0][0]

            self.balances = (self.cash_balance, self.bank_balance)
            print(f'committee_bank_balance: {self.bank_balance}')

            report_invoice_incomes = self.databaseConnection.query('select\n'
                                                                  'item_description,  sum(invoice_item_value)\n'
                                                                  ' from \n'
                                                                  ' invoice_receipt \n'
                                                                  ' join \n'
                                                                  ' invoice \n'
                                                                  ' on \n'
                                                                  ' invoice.invoice_no = invoice_receipt.invoice_no \n'
                                                                  ' join \n'
                                                                  ' invoice_line \n'
                                                                  ' on \n'
                                                                  ' invoice_line.invoice_no = invoice.invoice_no\n'
                                                                  ' join \n'
                                                                  ' item\n'
                                                                  ' on \n'
                                                                  ' item.item_code = invoice_line.item_code\n'
                                                                  ' where\n'
                                                                  'payment_datetime '
                                                                  f'between cast("{report_start_flip}" as DATETIME)'
                                                                  f' and cast("{report_end_flip}" as DATETIME)'
                                                                  'group by item.item_code;')
            report_incomes = self.databaseConnection.query('select\n'
                                                          'income_description,  '
                                                          'sum(cash_amount)+sum(transfer_amount)\n'
                                                          ' from \n'
                                                          ' income_receipt \n'
                                                          ' join \n'
                                                          ' income\n'
                                                          ' on \n'
                                                          ' income.income_id = income_receipt.income_id\n'
                                                          ' where\n'
                                                          f'payment_datetime between cast("{report_start_flip}" as DATETIME) '
                                                           f'and cast("{report_end_flip}" as DATETIME)'
                                                          'group by income.income_id;')
            report_incomes = report_invoice_incomes+report_incomes

            year_invoice_incomes = self.databaseConnection.query(f'select\n'
                                                                  'item_description,  sum(invoice_item_value)\n'
                                                                  ' from \n'
                                                                  ' invoice_receipt \n'
                                                                  ' join \n'
                                                                  ' invoice \n'
                                                                  ' on \n'
                                                                  ' invoice.invoice_no = invoice_receipt.invoice_no \n'
                                                                  ' join \n'
                                                                  ' invoice_line \n'
                                                                  ' on \n'
                                                                  ' invoice_line.invoice_no = invoice.invoice_no\n'
                                                                  ' join \n'
                                                                  ' item\n'
                                                                  ' on \n'
                                                                  ' item.item_code = invoice_line.item_code\n'
                                                                  ' where\n'
                                                                  'payment_datetime between date_format(now() , "%Y-01-01")'
                                                                  'and '
                                                                  f'"{report_end_flip}"\n'
                                                                  'group by item.item_code;')
            year_incomes = self.databaseConnection.query('select\n'
                                                          'income_description,  '
                                                          'sum(cash_amount)+sum(transfer_amount)\n'
                                                          ' from \n'
                                                          ' income_receipt \n'
                                                          ' join \n'
                                                          ' income\n'
                                                          ' on \n'
                                                          ' income.income_id = income_receipt.income_id\n'
                                                          ' where\n'
                                                          '(payment_datetime between date_format(now() , "%Y-01-01")'
                                                          ' and '
                                                          f'"{report_end_flip}") and income.income_id != 1 \n'
                                                          'group by income.income_id;')
            year_incomes = year_invoice_incomes+year_incomes

            report_expense = self.databaseConnection.query('select\n'
                                                          'expense_description,  '
                                                          'abs(sum(cash_amount)+sum(transfer_amount))\n'
                                                          ' from \n'
                                                          ' expense_receipt \n'
                                                          ' join \n'
                                                          ' expense\n'
                                                          ' on \n'
                                                          ' expense.expense_id = expense_receipt.expense_id\n'
                                                          ' where\n'
                                                          f'payment_datetime between cast("{report_start_flip}" as DATETIME) '
                                                          f'and cast("{report_end_flip}" as DATETIME) '
                                                          'group by expense.expense_id;')

            year_expense = self.databaseConnection.query('select\n'
                                                          'expense_description,  '
                                                          'abs(sum(cash_amount)+sum(transfer_amount))\n'
                                                          ' from \n'
                                                          ' expense_receipt \n'
                                                          ' join \n'
                                                          ' expense\n'
                                                          ' on \n'
                                                          ' expense.expense_id = expense_receipt.expense_id\n'
                                                          ' where\n'
                                                          'payment_datetime between date_format(now() , "%Y-01-01")'
                                                          ' and '
                                                          f'"{report_end_flip}" and expense.expense_id != 1\n'
                                                          'group by expense.expense_id;')
            unpaid_invoices = self.databaseConnection.query('select'
                                                 ' invoice.invoice_no,'
                                                 ' concat(member_fname," ", member_lname) as member_name,'
                                                 ' DATE_FORMAT(invoice_duedate,"%d-%m-%Y") as due_date,'
                                                 ' invoice_total '
                                                 'from invoice join members on invoice.member_no = members.member_no'
                                                 ' where invoice_no not in (select invoice_no from invoice_receipt) '
                                                 'order by invoice_no desc;')
            report_name =  f'Report({report_start.replace("/", "-")}_to_{report_end.replace("/", "-")}).pdf'

            report = CommiteeReportGenerator(report_name,
                                             datetime.today().date().strftime('%d/%m/%y'),
                                             report_start + ' to ' + report_end)

            report.incomes_list(report_incomes)
            report.expenses_list(report_expense)
            report.current_balances(self.balances)
            report.unpaid_invoices(unpaid_invoices)
            report.year_to_date_summary(year_incomes, year_expense)
            report.draw_image()
            report.save(f'{FOLDER}/committee_reports')
            self.destroy()
            view = messagebox.askyesno('Committee Report', 'Committee report was successfully created. Would you like to view it?' , parent=self.main_menu)
            if view:
                startfile(f'{FOLDER}\\committee_reports\\{report_name}')

        else:
            not_connected_message(self)


class BankDetails(tk.Tk):
    def __init__(self, connection, main_menu, *args, **kwargs):
        self.databaseConnection = connection
        self.main_menu = main_menu
        self.report_callback_exception = log_unhandled_exception
        tk.Tk.__init__(self, *args, **kwargs)
        self.title('Bank Details')
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self.grid_rowconfigure(2, weight=1)
        self.grid_rowconfigure(3, weight=1)

        x = self.winfo_screenwidth()
        y = self.winfo_screenheight()
        a = 0.35  # width
        b = 0.2  # height
        self.geometry(str(round(x * a)) +
                      'x'
                      + str(round(y * b)) +
                      '+'
                      + str(round((-a * x + x) / 2)) +
                      '+'
                      + str(round((-b * y + y) / 2)))

        self.bank_details = self.databaseConnection.query(f'select bank_name, bsb, account_no, customer_name from bank_detail')
        self.bank_details = self.bank_details[0]

        title_label = tk.Label(self, text='Bank Details')
        title_label.grid(row=0, columnspan=2)

        bsb_label = tk.Label(self, text='BSB:')
        bsb_label.grid(row=1, column=0)
        self.bsb_var = tk.StringVar(self)
        self.bsb_var.set(self.bank_details[1])
        bsb_entry = tk.Entry(self, textvariable=self.bsb_var, width=round(self.winfo_screenwidth() * 0.03))
        bsb_entry.grid(row=1, column=1)


        account_no_label = tk.Label(self, text='Account No:')
        account_no_label.grid(row=2, column=0)
        self.account_no_var = tk.StringVar(self)
        self.account_no_var.set(self.bank_details[2])
        account_no_entry = tk.Entry(self, textvariable=self.account_no_var, width=round(self.winfo_screenwidth() * 0.03))
        account_no_entry.grid(row=2, column=1)


        bank_name_label = tk.Label(self, text='Bank Name:')
        bank_name_label.grid(row=3, column=0)
        self.bank_name_var = tk.StringVar(self)
        self.bank_name_var.set(self.bank_details[0])
        bank_name_entry = tk.Entry(self, textvariable=self.bank_name_var, width=round(self.winfo_screenwidth() * 0.03))
        bank_name_entry.grid(row=3, column=1)

        customer_name_label = tk.Label(self, text='Customer Name:')
        customer_name_label.grid(row=4, column=0)
        self.customer_name_var = tk.StringVar(self)
        self.customer_name_var.set(self.bank_details[3])
        customer_name_entry = tk.Entry(self, textvariable=self.customer_name_var, width=round(self.winfo_screenwidth() * 0.03))
        customer_name_entry.grid(row=4, column=1)

        cancel = tk.Button(self, text='Cancel', width=6, command=self.destroy,
                           activebackground=button_pressed_color
                           )
        cancel.grid(column=0, row=5)
        cancel.bind('<Enter>', on_hover)
        cancel.bind('<Leave>', off_hover)


        submit = tk.Button(self, text='Submit', width=6, command=self.bank_detail_submit,
                           activebackground = button_pressed_color)
        submit.grid(column=1, row=5)
        submit.bind('<Enter>', on_hover)
        submit.bind('<Leave>', off_hover)

        self.attributes("-topmost", True)

    def bank_detail_submit(self):
        bsb = self.bsb_var.get()
        account_no = self.account_no_var.get()
        bank_name = self.bank_name_var.get()
        customer_name = self.customer_name_var.get()

        try:
            self.databaseConnection.insert(f'update bank_detail set '
                                           f'bank_name = "{bank_name}", '
                                           f'bsb = "{bsb}", '
                                           f'account_no = "{account_no}", '
                                           f'customer_name = "{customer_name}"')
            self.databaseConnection.commit()
        except Exception as e:
            self.databaseConnection.rollback()
            print('Error Logged')
            logger.exception(e)
            messagebox.showwarning('Unknown Error',
                                   f'An unknown error occurred and has been logged. Report to developer.')
        else:
            self.destroy()
            messagebox.showinfo('Bank Details', 'Bank details successfully updated.')


class EditCategories(tk.Tk):
    def __init__(self, connection, main_menu, *args, **kwargs):
        self.databaseConnection = connection
        self.main_menu = main_menu
        self.report_callback_exception = log_unhandled_exception
        tk.Tk.__init__(self, *args, **kwargs)
        self.title('Categories')
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self.grid_rowconfigure(2, weight=1)
        self.grid_rowconfigure(3, weight=1)

        y = self.winfo_screenheight()
        x = self.winfo_screenwidth()

        a = 400  # width
        b = 450  # height
        self.geometry(str(a) +
                      'x'
                      + str(b) +
                      '+'
                      + str(round((x-a)/2)) +
                      '+'
                      + str(round((y-b)/2)))


        title_label = tk.Label(self, text='Categories', font='Courier 15 bold')
        title_label.grid(row=0, columnspan=2)

        type_label = tk.Label(self, text='Type:')
        type_label.grid(row=1, column=0)

        self.categoryTypes = ['Expense', 'Income', 'Invoice']
        self.categoryTypeVar = tk.StringVar(self)
        self.categoryTypeVar.set(self.categoryTypes[0])
        self.categoryOptions = tk.OptionMenu(self, self.categoryTypeVar, *self.categoryTypes)
        self.categoryOptions.grid(row=1, column=1)

        self.table = Treeview(self, height=10)
        self.table.grid(row=2, columnspan=2)
        self.table["columns"] = ('Description', 'visibility')
        self.table.column('#0', width=40)
        self.table.heading('#0', text='ID')
        self.table.column('Description', width=250)
        self.table.heading('Description', text='Description')
        self.table.column('visibility', width=60)
        self.table.heading('visibility', text='Visibility')

        self.populate_table()

        self.categoryTypeVar.trace('w', self.populate_table)
        self.table.bind("<Double-1>", self.on_double_click)

        self.open_windows = []
        self.protocol("WM_DELETE_WINDOW", self.on_closing)


    def on_double_click(self, a):
        type = self.categoryTypeVar.get()
        item = self.table.selection()
        desc, visibility = self.table.item(item, 'value')
        id = self.table.item(item, 'text')
        EditCategoryRow(self.databaseConnection, (id,desc,visibility, type), self)


    def populate_table(self, *args):
        self.table.delete(*self.table.get_children())  # delete table rows
        self.expense_categories = self.databaseConnection.query(
            'select expense_id, expense_description, hidden from expense where expense_id != 1 order by expense_id;')
        self.income_categories = self.databaseConnection.query(
            'select income_id, income_description, hidden from income where income_id != 1 order by income_id;')
        self.invoice_categories = self.databaseConnection.query(
            'select item_code, item_description, hidden from item order by item_code;')
        if self.categoryTypeVar.get() == 'Expense':
            for row in self.expense_categories:
                desc = self.databaseConnection.sanitise(row[1])
                hidden = 'Hidden' if row[2] == 1 else 'Visible'
                self.table.insert("","end", text=row[0], values=[desc, hidden])
        elif self.categoryTypeVar.get() == 'Income':
            for row in self.income_categories:
                desc = self.databaseConnection.sanitise(row[1])
                hidden = 'Hidden' if row[2] == 1 else 'Visible'
                self.table.insert('', 'end', text=row[0], values=[desc, hidden])
        elif self.categoryTypeVar.get() == 'Invoice':
            for row in self.invoice_categories:
                desc = self.databaseConnection.sanitise(row[1])
                hidden = 'Hidden' if row[2] == 1 else 'Visible'
                self.table.insert('', 'end', text=row[0], values=[desc, hidden])

    def on_closing(self):
        for window in self.open_windows:
            window.destroy()
        self.destroy()


class EditCategoryRow(tk.Tk):
    def __init__(self, connection, row_info, Edit_main_window, *args, **kwargs):
        self.databaseConnection = connection
        self.edit_main_window = Edit_main_window
        self.edit_main_window.open_windows.append(self)
        self.report_callback_exception = log_unhandled_exception
        self.row = row_info
        tk.Tk.__init__(self, *args, **kwargs)
        self.title('Categories')
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        self.grid_columnconfigure(2, weight=1)
        self.grid_rowconfigure(0, weight=1)

        y = self.winfo_screenheight()
        x = self.winfo_screenwidth()

        a = 400  # width
        b = 75  # height
        self.geometry(str(a) +
                      'x'
                      + str(b) +
                      '+'
                      + str(round((x - a) / 2)) +
                      '+'
                      + str(round((y - b) / 2)))

        self.description_entry_var = tk.StringVar(self)
        self.description_entry_var.set(self.row[1])
        description_entry = tk.Entry(self, textvariable=self.description_entry_var)
        description_entry.grid(row=0, column=0)

        self.visibility_option_var = tk.StringVar(self)
        self.visibility_option_var.set(self.row[2])
        visibility_option = tk.OptionMenu(self, self.visibility_option_var, *['Visible', 'Hidden'])
        visibility_option.grid(row=0, column=1)

        save_button = tk.Button(self, text='Save', command=self.save_row)
        save_button.grid(row=0, column=2)

        cancel_button = tk.Button(self, text='Cancel', command=self.on_closing)
        cancel_button.grid(row=0, column=3)

        self.protocol("WM_DELETE_WINDOW", self.on_closing)

    def on_closing(self):
        self.edit_main_window.open_windows.remove(self)
        self.destroy()

    def save_row(self):
        desc = self.description_entry_var.get()
        visibility = self.visibility_option_var.get()
        if visibility == 'Visible':
            hidden = 0
        else:
            hidden = 1
        id = self.row[0]
        type = self.row[3]
        if type == 'Expense':
            table = 'expense'
            id_column = 'expense_id'
        elif type == 'Income':
            table = 'income'
            id_column = 'income_id'
        elif type == 'Invoice':
            table = 'item'
            id_column = 'item_code'
        try:
            self.databaseConnection.insert(f'update {table} set '
                                           f'{table}_description = "{self.databaseConnection.sanitise(desc)}", '
                                           f'hidden = "{hidden}" where {id_column} = {id};')
            self.databaseConnection.commit()
        except Exception as e:
            self.databaseConnection.rollback()
            print('Error Logged')
            logger.exception(e)
            messagebox.showwarning('Unknown Error',
                                   f'An unknown error occurred and has been logged. Report to developer.', parent=self.edit_main_window)
        else:
            self.on_closing()
            self.edit_main_window.populate_table()

        print('save')


logger = logging.getLogger(__name__)
logging.basicConfig(filename=f'{FOLDER}/log_file.txt',
                    format='-' * 60 + '\n%(asctime)s - %(name)s - %(levelname)s - %(message)s')
# logger.setLevel(logging.DEBUG)

import webbrowser


if TOKEN == '':
    url = f'https://accounts.google.com/o/oauth2/v2/auth?scope={scope}&response_type=code&redirect_uri={redirect_uri}&client_id={CLIENT_ID}'
    webbrowser.open(url)
else:



    db_connection_error = 0
    if DATABASE_NAME != '' and DATEBASE_PASSWORD != '' and DATABASE_USER != '' and DATABASE_HOST != '':
        try:
            print(f'Connecting to {DATABASE_USER}@{DATABASE_HOST}..................', end='')
            databaseConnection = connect_database(DATABASE_USER, DATEBASE_PASSWORD, DATABASE_HOST, DATABASE_NAME)
        except Exception:
            print('Unknown connection Issue.')
            databaseConnection = None
            db_connection_error = 1
    else:
        databaseConnection = None






    app = App(EMAIL_ADDRESS, EMAIL_PASSWORD, EMAIL_HOST, EMAIL_PORT, connection=databaseConnection, connection_error=db_connection_error)  # initialise app
    print("*"*30)
    print('App loaded')
    app.protocol("WM_DELETE_WINDOW", on_closing)
    app.mainloop()

    # db_errors().ProgrammingError ##wrong password, wrong_user, wrong database
    # db_errors().InterfaceError  ## wrong host
